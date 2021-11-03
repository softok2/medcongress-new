import base64
import json
from datetime import datetime, timedelta
from os import remove
from pathlib import Path

import openpyxl
import pandas as pd
from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin
from django.core.exceptions import RequestDataTooBig, ValidationError
from django.core.mail import EmailMessage
from django.db import connections
from django.db.models import Q, Sum
from django.http import (HttpResponse, HttpResponseBadRequest,
                         HttpResponseRedirect, JsonResponse)
from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.utils.crypto import get_random_string
from django.views.generic import ListView, TemplateView
from django.views.generic.edit import (CreateView, DeleteView, FormView,
                                       UpdateView)
from MedCongressAdmin.apps import validarOrganizador, validarUser
from MedCongressAdmin.forms.congres_forms import (AsignarCongresoForms,
                                                  AsignarConstanciaUserForms,
                                                  CongresoCategPagoForm,
                                                  CongresoForms,
                                                  CongresoPatrocinadorForm,
                                                  CongresoProgramaForm,
                                                  CongresoSalaForm,
                                                  CongresoSocioForm,
                                                  ExportarExelForm,
                                                  ImagenCongForms,
                                                  PonenciaForms,
                                                  ExportarLogsCongresoExelForm,
                                                  AsignarConstanciasForm)
from MedCongressAdmin.task import AsignarBeca, Constancia
from MedCongressApp.models import (AvalCongreso, Bloque, CategoriaPagoCongreso,
                                   Congreso, CuestionarioPregunta,
                                   CuestionarioRespuestas, DocumentoPrograma,
                                   ImagenCongreso, Moderador, Organizador,
                                   PerfilUsuario, Ponencia, Ponente,
                                   PreguntasFrecuentes, RelCongresoAval,
                                   RelCongresoCategoriaPago, RelCongresoSocio,
                                   RelCongresoUser, RelPonenciaPonente,
                                   RelTalleresCategoriaPago, RelTallerUser,
                                   Sala, SocioCongreso, Taller,
                                   TrabajosInvestigacion, Ubicacion, User,
                                   UserActivityLog,ConstanciaUsuario)
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Protection, Side)
from pandas import isnull
from PIL import Image, ImageDraw, ImageFont


class CongressListView(validarOrganizador,ListView):
    model = Congreso
    context_object_name = 'congress'
    template_name = 'MedCongressAdmin/congress/listar.html'
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        if self.request.GET.get('search'):
            context['search']=self.request.GET.get('search')
        context['congress']=Congreso.objects.all()
        return context

class CongressCreateView(validarUser,FormView):
    form_class = CongresoForms
    success_url = reverse_lazy('MedCongressAdmin:congress_list')
    template_name = 'MedCongressAdmin/congress/form.html'

    def form_valid(self, form):
        try:
            congress=form['congreso'].save(commit=False)
            # imagen=form['imagen_congreso'].save(commit=False)
            ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)
            if ubic.exists():
                congress.lugar=ubic.first()
            else:
                ubicacion=form['ubicacion'].save(commit=True)
                congress.lugar=ubicacion    
            path=congress.titulo.replace("/","").replace(" ","-").replace("?","").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ñ","n")
            congress.path=path  
            
            image_64_encode=self.request.POST['congreso-prueba1']
            campo = image_64_encode.split(",")
            image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
            chars = '0123456789'
            nombre = get_random_string(5, chars)
            image_result = open('MedCongressApp/static/congreso/imagen_seg_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
            image_result.write(image_64_decode)
            congress.imagen_seg='congreso/imagen_seg_%s.png'%(nombre)

            image_64_encode=self.request.POST['congreso-imagen_home']
            campo = image_64_encode.split(",")
            image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
           
            chars = '0123456789'
            nombre = get_random_string(5, chars)
            image_result = open('MedCongressApp/static/congreso/imagen_home_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
            image_result.write(image_64_decode)
            congress.imagen_home='congreso/imagen_home_%s.png'%(nombre)

            if self.request.POST['congreso-constancia']:
                image_64_encode=self.request.POST['congreso-constancia']
                campo = image_64_encode.split(",")
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_result = open('MedCongressApp/static/congreso/img_constancia/imagen_constancia_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                congress.foto_constancia='congreso/img_constancia/imagen_constancia_%s.png'%(nombre)
            if self.request.POST['congreso-const_moderador']:
                image_64_encode=self.request.POST['congreso-const_moderador']
                campo = image_64_encode.split(",")
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_result = open('MedCongressApp/static/congreso/img_constancia/imagen_const_moderador_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                congress.foto_const_moderador='congreso/img_constancia/imagen_const_moderador_%s.png'%(nombre)
            if self.request.POST['congreso-const_ponente']:
                image_64_encode=self.request.POST['congreso-const_ponente']
                campo = image_64_encode.split(",")
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_result = open('MedCongressApp/static/congreso/img_constancia/imagen_const_ponente_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                congress.foto_const_ponente='congreso/img_constancia/imagen_const_ponente_%s.png'%(nombre)
            congress.is_home=False
            congress.save()
            cant=0
            for sala in self.request.POST.getlist('salas'):
                resp=Sala(congreso=congress,  titulo=self.request.POST.getlist('salas')[cant],published=True)
                resp.save() 
                cant=cant+1   
            for respuesta in self.request.POST.getlist('congreso-prueba'):
                image_64_encode=respuesta
                campo = image_64_encode.split(",")
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_result = open('MedCongressApp/static/congreso/imagen_programa_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                campo_imagen='congreso/imagen_programa_%s.png'%(nombre)
                imagen=ImagenCongreso(congreso=congress,imagen=campo_imagen)
                imagen.save()
           

            
            return super().form_valid(form)
        except Exception as e:
            messages.warning(self.request, e)
            return super().form_invalid(form)

    def get_success_url(self):
        url =  reverse_lazy('MedCongressAdmin:congress_list')
        if self.request.GET.get('search'):
            self.success_url =  '%s?search=%s'%(url,self.request.GET.get('search'))
        else:
            self.success_url =  url
        return self.success_url

class CongressUpdateView(validarOrganizador,FormView):
    form_class = CongresoForms
    success_url = reverse_lazy('MedCongressAdmin:congress_list')
    template_name = 'MedCongressAdmin/congress/form.html'

    def get(self, request, **kwargs):
        congreso=Congreso.objects.filter(pk=self.kwargs.get('pk')).first()
        if congreso is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        return self.render_to_response(self.get_context_data()) 

    def get_queryset(self, **kwargs):
        return Congreso.objects.filter(pk=self.kwargs.get('pk'))
    
    def get_form_kwargs(self):
        kwargs = super(CongressUpdateView, self).get_form_kwargs()
        self.object=Congreso.objects.get(pk=self.kwargs.get('pk'))
        kwargs.update(instance={
            'congreso': self.object,
            'ubicacion': self.object.lugar,
            
        })
        return kwargs

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        self.object=Congreso.objects.get(pk=self.kwargs.get('pk'))
        imagenes=ImagenCongreso.objects.filter(congreso=self.object)
        context['salas']=Sala.objects.filter(congreso=self.object)
        context['update']=self.object
        if imagenes:
            context['imagenes']=imagenes
        if self.object.imagen_seg:
            context['imagen_seg_url']=self.object.imagen_seg
        if self.object.meta_og_imagen:
            context['imagen_meta']='/static/%s'%(self.object.meta_og_imagen)
        if self.object.foto_constancia:
            context['foto_constancia']='/static/%s'%(self.object.foto_constancia)
        if self.object.foto_const_ponente:
            context['foto_const_ponente']='/static/%s'%(self.object.foto_const_ponente)
        if self.object.foto_const_moderador:
            context['foto_const_moderador']='/static/%s'%(self.object.foto_const_moderador)    
        if self.object.imagen_home:
            context['imagen_home']=self.object.imagen_home
        return context   
    
    def form_valid(self, form):
        update_congreso=Congreso.objects.get(pk=self.request.POST['update']) 
        try:
            congress=form['congreso'].save(commit=False)
            # imagen=form['imagen_congreso'].save(commit=False)
            ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)
            if ubic.exists():
                congress.lugar=ubic.first()
            else:
                ubicacion=Ubicacion(direccion=form['ubicacion'].instance.direccion,latitud=form['ubicacion'].instance.latitud,longitud=form['ubicacion'].instance.longitud)
                ubicacion.save()
                congress.lugar=ubicacion

            imagen_seg=self.request.POST['congreso-prueba1']
            if 'congreso/' not in imagen_seg:
            
                image_64_encode=self.request.POST['congreso-prueba1']
                campo = image_64_encode.split(",")
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                image_result = open('MedCongressApp/static/congreso/imagen_seg_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                if  update_congreso.imagen_seg:
                    fileObj = Path('MedCongressApp/static/%s'%( update_congreso.imagen_seg))
                    if fileObj.is_file():
                        remove('MedCongressApp/static/%s'%( update_congreso.imagen_seg))
                congress.imagen_seg='congreso/imagen_seg_%s.png'%(nombre)
            ###### Imagen Principal
            imagen_home=self.request.POST['congreso-imagen_home']
            if 'congreso/' not in imagen_home:
                
                image_64_encode=self.request.POST['congreso-imagen_home']
                campo = image_64_encode.split(",")
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                image_result = open('MedCongressApp/static/congreso/imagen_home_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                if  update_congreso.imagen_home:
                    fileObj = Path('MedCongressApp/static/%s'%( update_congreso.imagen_home))
                    if fileObj.is_file():
                        remove('MedCongressApp/static/%s'%( update_congreso.imagen_home))
                congress.imagen_home='congreso/imagen_home_%s.png'%(nombre)
            ###### Imagen Constancia para Moderadores
            if self.request.POST['congreso-const_moderador']:
                imagen_const_moderador=self.request.POST['congreso-const_moderador']
                if 'congreso/' not in imagen_const_moderador:
                    
                    image_64_encode=self.request.POST['congreso-const_moderador']
                    campo = image_64_encode.split(",")
                    chars = '0123456789'
                    nombre = get_random_string(5, chars)
                    image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                    image_result = open('MedCongressApp/static/congreso/imagen_const_moderador_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                    image_result.write(image_64_decode)
                    if  update_congreso.foto_const_moderador:
                        fileObj = Path('MedCongressApp/static/%s'%( update_congreso.foto_const_moderador))
                        if fileObj.is_file():
                            remove('MedCongressApp/static/%s'%( update_congreso.foto_const_moderador))
                    congress.foto_const_moderador='congreso/imagen_const_moderador_%s.png'%(nombre)
            ###### Imagen Constancia para Ponentes
            if self.request.POST['congreso-const_ponente']:
                imagen_const_ponente=self.request.POST['congreso-const_ponente']
                if 'congreso/' not in imagen_const_ponente:
                    
                    image_64_encode=self.request.POST['congreso-const_ponente']
                    campo = image_64_encode.split(",")
                    chars = '0123456789'
                    nombre = get_random_string(5, chars)
                    image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                    image_result = open('MedCongressApp/static/congreso/imagen_const_ponente_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                    image_result.write(image_64_decode)
                    if  update_congreso.foto_const_ponente:
                        fileObj = Path('MedCongressApp/static/%s'%( update_congreso.foto_const_ponente))
                        if fileObj.is_file():
                            remove('MedCongressApp/static/%s'%( update_congreso.foto_const_ponente))
                    congress.foto_const_ponente='congreso/imagen_const_ponente_%s.png'%(nombre)

            if self.request.POST['congreso-constancia']:
                constancia=self.request.POST['congreso-constancia']
                if 'congreso/' not in constancia:
                
                    image_64_encode=self.request.POST['congreso-constancia']
                    campo = image_64_encode.split(",")
                    chars = '0123456789'
                    nombre = get_random_string(5, chars)
                    image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                    image_result = open('MedCongressApp/static/congreso/img_constancia/imagen_constancia_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                    image_result.write(image_64_decode)
                    if  update_congreso.foto_constancia:
                        fileObj = Path('MedCongressApp/static/%s'%( update_congreso.foto_constancia))
                        if fileObj.is_file():
                            remove('MedCongressApp/static/%s'%( update_congreso.foto_constancia))
                    congress.foto_constancia='congreso/img_constancia/imagen_constancia_%s.png'%(nombre)    
            update_congreso=congress
            update_congreso.save()
            ImagenCongreso.objects.filter(congreso=update_congreso).delete()
            for respuesta in self.request.POST.getlist('congreso-prueba'):
                if 'congreso/' in respuesta:
                    imagen=ImagenCongreso(congreso=update_congreso,imagen=respuesta)
                    imagen.save()
                else:
                    image_64_encode=respuesta
                    campo = image_64_encode.split(",")
                    image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                    chars = '0123456789'
                    nombre = get_random_string(5, chars)
                    image_result = open('MedCongressApp/static/congreso/imagen_programa_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                    image_result.write(image_64_decode)
                    campo_imagen='congreso/imagen_programa_%s.png'%(nombre)
                    imagen=ImagenCongreso(congreso=update_congreso,imagen=campo_imagen)
                    imagen.save()
            cant=0
            for sala in self.request.POST.getlist('salas'):
                resp=Sala.objects.get(id=self.request.POST.getlist('salas_id')[cant])
                resp.titulo=self.request.POST.getlist('salas')[cant]
                resp.save() 
                cant=cant+1     
            return super().form_valid(form)
        except RequestDataTooBig as e:
            messages.warning(self.request, e)
            return super().form_invalid(form)
    
    def get_success_url(self):
        url =  reverse_lazy('MedCongressAdmin:congress_list')
        if self.request.GET.get('search'):
            self.success_url =  '%s?search=%s'%(url,self.request.GET.get('search'))
        else:
            self.success_url =  url
        return self.success_url

class AddPonenciaCongreso(validarUser,TemplateView):
    def get(self, request):
        
        if request.is_ajax:
            id_ponencia =request.GET.get("id_ponencia")
            congreso_path =request.GET.get("congreso")
            
            congreso=Congreso.objects.get(path=congreso_path)
            ponencia=Ponencia.objects.get(id=id_ponencia)
            ponencia.congreso=congreso
            ponencia.save()
            return JsonResponse({'succes':True}, safe=False)
        return TemplateResponse(request, reverse('home'))

class CongressDeletedView(validarUser,DeleteView):
    model = Congreso
    success_url = reverse_lazy('MedCongressAdmin:congress_list')
    # template_name = 'MedCongressAdmin/country_form.html'
  
class  CongressPonenteCreateView(validarUser,CreateView):
   
    form_class = CongresoCategPagoForm
    # success_url = reverse_lazy('MedCongressAdmin:ponencias_list')
    template_name = 'MedCongressAdmin/congreso_ponencia_form.html'
    def form_valid(self, form):
        ponencia=form.save(commit=False)
        ponencia.save()
        return super(CongressPonenteCreateView, self).form_valid(form)

    def get_success_url(self):
           self.success_url =  reverse_lazy('MedCongressAdmin:Congres_pagos',kwargs={'id': self.kwargs.get('path')} )
           return self.success_url

    def get_context_data(self, **kwargs):
        ctx = super(CongressCategPagosCreateView, self).get_context_data(**kwargs)
        pon=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        ctx['cong'] = pon
        ctx['sala']= Sala.objects.filter(congreso=pon)
        ctx['bloque']= Bloque.objects.filter(congreso=pon)
        
        return ctx

def GetBloques(request):
    data = json.dumps([])
    if request.is_ajax():
        query = request.POST['congreso_id']
        bloques=Bloque.objects.filter(congreso=Congreso.objects.get(pk=query))
        results = []
        for bloque in bloques:
            results.append({'titulo':bloque.titulo,'id':bloque.pk})
        data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

def GetSalas(request):
    data = json.dumps([])
    if request.is_ajax():
        query = request.POST['congreso_id']
        bloques=Sala.objects.filter(congreso=Congreso.objects.get(pk=query))
        results = []
        for bloque in bloques:
            results.append({'titulo':bloque.titulo,'id':bloque.pk})
        data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

def GetPagos(request):
    if request.is_ajax():
        query = request.POST['congreso_id']
        categoria=RelCongresoCategoriaPago.objects.filter(congreso=Congreso.objects.get(pk=query))
       
        results = []
        for cat in categoria:
            results.append({'nombre':cat.categoria.nombre,'id':cat.categoria.pk,'moneda':cat.moneda})
            data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)   

class AsignarCongressListView(validarOrganizador,ListView,FormView):
    model = RelCongresoUser
    context_object_name = 'congress'
    template_name = 'MedCongressAdmin/congress/listar_asignar_congreso.html'
    form_class=ExportarExelForm
    
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs(*args, **kwargs)
        kwargs['user'] =self.request.user
        return kwargs

    def form_valid(self, form):
        self.object_list = self.get_queryset()
        id_congreso=self.request.POST['congreso']
        query= RelCongresoUser.objects.filter(congreso=id_congreso,is_pagado=True).values('user__usuario__first_name','user__usuario__last_name','user__usuario__email','user__genero__denominacion','categoria_pago__nombre','user__cel_profecional','user__categoria__nombre','user__ubicacion__direccion','user__especialidad__nombre','user__fecha_nacimiento','user__num_telefono','congreso__titulo','is_beca','user__puesto').annotate(Sum('cantidad'))
        
        if query:
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=40
            ws.column_dimensions['C'].width=40
            ws.column_dimensions['D'].width=47

            ws.column_dimensions['E'].width=20
            ws.column_dimensions['F'].width=20
            ws.column_dimensions['G'].width=27
            ws.column_dimensions['H'].width=56
            ws.column_dimensions['I'].width=20
            ws.column_dimensions['J'].width=25
            ws.column_dimensions['K'].width=25
            ws.column_dimensions['L'].width=22
            ws.column_dimensions['N'].width=50

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            celdas_beca = NamedStyle(name="celdas_beca")
            celdas_beca.font=Font(size=12,color='00CCCC')
            
            celdas_beca.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas_beca.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Usuarios que han comprado el  Congresos :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            ws['A2'] ='" %s "'%(query[0]['congreso__titulo']) 
            ws['A2'].font = Font(size=12,bold=True)
            ws['A2'].alignment = Alignment(mergeCell='center',horizontal='center') 
           
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'].style =titulo
            ws['B3'].style =titulo 
            ws['C3'].style =titulo 
            ws['D3'].style =titulo                   
            ws['E3'].style =titulo 
            ws['F3'].style =titulo 
            ws['G3'].style =titulo 
            ws['H3'].style =titulo
            ws['I3'].style =titulo 
            ws['J3'].style =titulo  
            ws['K3'].style =titulo 
            ws['L3'].style =titulo
            ws['M3'].style =titulo
            ws['N3'].style =titulo

            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Dirección'
            ws['E3'] = 'Teléfono'
            ws['F3'] = 'Género'
            ws['G3'] = 'Categoría'
            ws['H3'] = 'Especialidad'
            ws['I3'] = 'Cédula Profecional' 
            ws['J3'] = 'Fecha de Nacimiento' 
            ws['K3'] = 'Categoría de Pago'  
            ws['L3'] = 'Cantidad Comprados' 
            ws['M3'] = 'Beca' 
            ws['N3'] = 'Lugar de Trabajo'             
            cont=4
            
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                
                ws.cell(row=cont,column=1).style=celdas
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).style=celdas
                ws.cell(row=cont,column=2).value ='%s %s'%(quer['user__usuario__first_name'],quer['user__usuario__last_name']) 
                ws.cell(row=cont,column=3).style=celdas
                ws.cell(row=cont,column=3).value = quer['user__usuario__email']
                ws.cell(row=cont,column=4).style=celdas
                ws.cell(row=cont,column=4).value = quer['user__ubicacion__direccion']
                ws.cell(row=cont,column=5).style=celdas
                ws.cell(row=cont,column=5).value = quer['user__num_telefono']
                ws.cell(row=cont,column=6).style=celdas
                ws.cell(row=cont,column=6).value = quer['user__genero__denominacion']
                ws.cell(row=cont,column=7).style=celdas
                ws.cell(row=cont,column=7).value = quer['user__categoria__nombre']
                ws.cell(row=cont,column=8).style=celdas
                ws.cell(row=cont,column=8).value = quer['user__especialidad__nombre']
                ws.cell(row=cont,column=9).style=celdas
                ws.cell(row=cont,column=9).value = quer['user__cel_profecional']
                ws.cell(row=cont,column=10).style=celdas
                ws.cell(row=cont,column=10).value = quer['user__fecha_nacimiento']
                ws.cell(row=cont,column=11).style=celdas
                ws.cell(row=cont,column=11).value = quer['categoria_pago__nombre']
                ws.cell(row=cont,column=12).style=celdas
                ws.cell(row=cont,column=12).value = quer['cantidad__sum']
                beca='No'
                ws.cell(row=cont,column=13).style=celdas
                if quer['is_beca']:
                    beca='Si'
                    ws.cell(row=cont,column=13).style=celdas_beca
                ws.cell(row=cont,column=13).value = beca
                ws.cell(row=cont,column=14).style=celdas
                ws.cell(row=cont,column=14).value = quer['user__puesto']
                cont = cont + 1
                
            
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=RelCongresoUser.xlsx"
            wb.save(response)
            return response
        else:
            congreso=Congreso.objects.get(pk=id_congreso)
            messages.warning(self.request, 'Todavía ningún usuario ha comprado este congreso')
            return HttpResponseRedirect(reverse_lazy('MedCongressAdmin:asig_congress_list')+'?exportar=%s'%(congreso.path))

    def get_context_data(self, **kwargs):
        context = super(AsignarCongressListView, self).get_context_data(**kwargs)
        context['search']=self.request.GET.get('search')
        if self.request.GET.get('exportar'):
            congreso_evn=[]
            congresos= Congreso.objects.all()
            activo=False
            for congreso in congresos:
                if congreso.path == self.request.GET.get('exportar'):
                    activo=True
                else:
                    activo=False
                congreso_evn.append({ 'id':congreso.pk,
                                    'titulo':congreso.titulo,
                                        'activo':activo,

                })
            context['exportar']= congreso_evn
            
        return context  
    
class AsignarCongressAddViews(validarOrganizador,FormView):
    form_class = AsignarCongresoForms
    success_url = reverse_lazy('MedCongressAdmin:asig_congress_list')
    template_name = 'MedCongressAdmin/congress/asig_congress_form.html'

    def form_valid(self, form):
        congress=form.save(commit=True)
        
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super(AsignarCongressAddViews, self).get_context_data(**kwargs)
        ctx['search']=self.request.GET.get('search')

        if self.request.GET.get('ponente'):
             ctx['ponente']=self.request.GET.get('ponente')
        if self.kwargs.get('pk'):
            usuario=PerfilUsuario.objects.get(pk=self.kwargs.get('pk'))
            ctx['usuario'] = usuario
        return ctx 
    
    def get_success_url(self):
        if self.request.GET.get('ponente'):
            url=reverse_lazy('MedCongressAdmin:usuarios_list') 
        else:
            url=reverse_lazy('MedCongressAdmin:asig_congress_list')
        self.success_url='%s?search=%s'%(url,self.request.GET.get('search'))
        
        return self.success_url          

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs(*args, **kwargs)
        kwargs['user'] =self.request.user
        return kwargs 

class AsignarCongressDeletedViews(validarOrganizador,DeleteView):
    model = RelCongresoUser
    success_url = reverse_lazy('MedCongressAdmin:asig_congress_list')
    
class CongressImagenDeletedView(validarUser,DeleteView):
    model = ImagenCongreso
    success_url = reverse_lazy('MedCongressAdmin:cat_usuarios_list')

class Ver_usuarios (validarUser,TemplateView):

    template_name='MedCongressAdmin/ver_usuarios.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        congresos=Congreso.objects.all()
        cat_pago=CategoriaPagoCongreso.objects.all()
        context['congresos']=congresos
        context['cat_pagos']=cat_pago
       
        return context

    def post(self, request, **kwargs):

        id_congreso=self.request.POST['congresos']
        id_cat_pago=self.request.POST['cat_pago']
        cat_pago= CategoriaPagoCongreso.objects.get(pk=id_cat_pago)
        congreso= Congreso.objects.get(pk=id_congreso)
        archivo_excel = pd.read_excel(self.request.FILES['exel'])
        values = archivo_excel['Correo'].values
        ###################
        sin_pagar=[]
        for value in values:
            user=User.objects.filter(email=value).first()
            if PerfilUsuario.objects.filter(usuario=user).exists():
                if user :
                    relacion=RelCongresoUser.objects.filter(congreso=congreso,user=user.perfilusuario).first()
                    if relacion:
                        relacion.is_pagado=True
                        relacion.cantidad=1
                        relacion.save()
                    else:
                        rel=RelCongresoUser(congreso=congreso,user=user.perfilusuario,is_pagado=True,cantidad=1,categoria_pago=cat_pago)
                        rel.save()
                else:
                    sin_pagar.append(value)
            else:
                sin_pagar.append(value)
        data = {'Usuarios que no se han Autentificado': sin_pagar}
        df = pd.DataFrame(data, columns = ['Usuarios que no se han Autentificado'])
        df.to_excel('MedCongressApp/static/patrocinadores/example.xlsx', sheet_name='example')
       
        ###################

        return HttpResponseRedirect(reverse('MedCongressAdmin:Ver_exel' ))

class Ver_Exel(validarUser,TemplateView):

    template_name='MedCongressAdmin/ver_exel.html'

class Exportar_usuarios(validarUser,TemplateView):
    template_name='MedCongressAdmin/view_exportar_usuario.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        usuarios=  PerfilUsuario.objects.all()
        email=[]
        nombre=[]
        for usuario in usuarios:
            if Ponente.objects.filter(user=usuario).exists():
                email=email
            else:
                email.append(usuario.usuario.email)
                nombre.append('%s %s'%(usuario.usuario.first_name,usuario.usuario.last_name))
        data = {'Nombre y Apellidos':nombre,'Email': email}
        df = pd.DataFrame(data, columns = ['Nombre y Apellidos','Email'])
        df.to_excel('MedCongressApp/static/patrocinadores/user_registrados.xlsx', sheet_name='example')
        return context

class Usuarios_pagaron(validarUser,TemplateView):
    template_name='MedCongressAdmin/view_pagaron_usuario.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        usurios_pagaron= RelCongresoUser.objects.filter(is_pagado=True).distinct('user')
        email=[]
        nombre=[]
        for usuario in usurios_pagaron:
                email.append(usuario.user.usuario.email)
                nombre.append('%s %s'%(usuario.user.usuario.first_name,usuario.user.usuario.last_name))
        data = {'Nombre y Apellidos':nombre,'Email': email}
        df = pd.DataFrame(data, columns = ['Nombre y Apellidos','Email'])
        df.to_excel('MedCongressApp/static/patrocinadores/user_pagaron.xlsx', sheet_name='example')
        return context

    def post(self, request):
        #Obtenemos todas las personas de nuestra base de datos
        congreso=self.request.POST['congreso']
        query= RelCongresoUser.objects.filter(congreso=congreso,is_pagado=True).values('user__usuario__first_name','user__usuario__last_name','user__usuario__email','congreso__titulo','categoria_pago__nombre','user__ubicacion__direccion','beca').annotate(Sum('cantidad'))

		#Creamos el libro de trabajo
        wb = Workbook()
		#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
       
		#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        if query:
            ws['B1'] = 'Usuarios que han comprado el Congresos %s'%(query[0]['congreso__titulo'])
            ws['B1'].font = Font(size=12,bold=True)
            ws['B1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Congreso'
            ws['E3'] = 'Categoria de Pago'
            ws['F3'] = 'Cantidad'
            ws['G3'] = 'Dirección'
            ws['H3'] = 'Beca'           
            cont=4
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).value = '%s %s'%(quer['user__usuario__first_name'],quer['user__usuario__last_name'])
                ws.cell(row=cont,column=3).value = quer['user__usuario__email']
                ws.cell(row=cont,column=4).value = quer['congreso__titulo']
                ws.cell(row=cont,column=5).value = quer['categoria_pago__nombre']
                ws.cell(row=cont,column=6).value = quer['cantidad__sum']
                ws.cell(row=cont,column=6).value = quer['user__ubicacion__direccion']
                beca='No'
                if quer['beca']:
                    beca='Si'
                ws.cell(row=cont,column=7).value = beca
                cont = cont + 1
        
           
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=RelCongresoUser.xlsx"
            wb.save(response)
            return response
        else:
            return self.render_to_response(self.get_context_data())
 
class  PatrocinadorSeleccionarView(validarOrganizador,FormView):
    
    model=RelCongresoAval
    form_class = CongresoPatrocinadorForm
    # success_url = reverse_lazy('MedCongressAdmin:ponencias_list')
    template_name='MedCongressAdmin/congress/patrocinador_form.html'
   
    def get(self, request, **kwargs):
        congreso=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        if congreso is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403')) 
        return self.render_to_response(self.get_context_data()) 

    def get_success_url(self):
           self.success_url =  reverse_lazy('MedCongressAdmin:patrocinadores_list' )+'?congreso='+self.kwargs.get('path')
           return self.success_url

    def get_context_data(self, **kwargs):
        context = super(PatrocinadorSeleccionarView, self).get_context_data(**kwargs)
        pon=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        context['congreso'] = pon
        return context

    def form_valid(self, form):
        relacion_aval=form.save(commit=True)
        return super().form_valid(form)

def PatrocinadorSeleccionarDeleted( request):
        
    if request.is_ajax():
        id_aval =request.GET.get("pk")
        patrocinador=AvalCongreso.objects.get(pk=id_aval)
        id_congreso=request.GET.get("pk_congreso")
        congreso=Congreso.objects.get(pk=id_congreso)
        relacion= RelCongresoAval.objects.filter(congreso=congreso,aval=patrocinador)
       
        if relacion.exists():
            relacion.first().delete()
            return JsonResponse({'success':True}, safe=False)
    return JsonResponse({'success':False}, safe=False)

def SocioSeleccionarDeleted( request):
        
    if request.is_ajax():
        id_socio =request.GET.get("pk")
        socio=SocioCongreso.objects.get(pk=id_socio)
        id_congreso=request.GET.get("pk_congreso")
        congreso=Congreso.objects.get(pk=id_congreso)
        relacion= RelCongresoSocio.objects.filter(congreso=congreso,socio=socio)
        if relacion.exists():
            relacion.first().delete()
            return JsonResponse({'success':True}, safe=False)
    return JsonResponse({'success':False}, safe=False)

class  SocioSeleccionarView(validarOrganizador,FormView):
    
    model=RelCongresoSocio
    form_class = CongresoSocioForm
    # success_url = reverse_lazy('MedCongressAdmin:ponencias_list')
    template_name='MedCongressAdmin/congress/socio_form.html'
   
    def get(self, request, **kwargs):
        congreso=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        if congreso is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403')) 
        return self.render_to_response(self.get_context_data()) 
             
    def get_success_url(self):
        self.success_url =  reverse_lazy('MedCongressAdmin:socios_list' )+'?congreso='+self.kwargs.get('path')
        return self.success_url

    def get_context_data(self, **kwargs):
        context = super(SocioSeleccionarView, self).get_context_data(**kwargs)
        
        pon=Congreso.objects.filter(path=self.kwargs.get('path')).first()
       
        context['congreso'] = pon
        return context
    
    def form_valid(self, form):
    
        relacion_aval=form.save(commit=True)
       
        return super().form_valid(form)

class CongresoDetail(TemplateView):
    # template_name= 'MedCongressApp/congreso_detail.html' 
    

    def get(self, request, **kwargs):
       
        congreso=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        if congreso is None:
            return   HttpResponseRedirect(reverse('Error404'))
        return self.render_to_response(self.get_context_data())

    def get_template_names(self):
        congreso=Congreso.objects.filter(path=self.kwargs.get('path')).first()
        if congreso.is_openpay:
            template_name= 'MedCongressApp/congreso_detail_openpay.html'
        else:
            template_name= 'MedCongressApp/%s'%(congreso.template)
        return template_name

    def get_context_data(self, **kwargs):
        context = super(CongresoDetail, self).get_context_data(**kwargs)
        congreso=Congreso.objects.filter(path=self.kwargs.get('path')).first()
       
        context['patrocinadores']=RelCongresoAval.objects.filter(congreso=congreso)
        context['socios']=RelCongresoSocio.objects.filter(congreso=congreso)
        if self.request.user.is_authenticated:
            pagado=RelCongresoUser.objects.filter(congreso=congreso,user=self.request.user.perfilusuario,is_pagado=True)
            if pagado :
                context['pagado']=True
                # InsertLog(congreso.pk,'Congreso',self.request.user.perfilusuario)
                constancias=RelCongresoUser.objects.filter(congreso=congreso,user=self.request.user.perfilusuario)
                for constancia in constancias:
                    if constancia.is_constancia:
                        context['constancia']=True
        salas=Sala.objects.filter(congreso=congreso,published=True).exclude( cod_video__isnull=True).exclude(cod_video__exact='') 
       
        salas_env=[]
        for sala in salas:
            
            if sala.ponencia_streamming:
                ponencia_sala=Ponencia.objects.get(pk=sala.ponencia_streamming)
            else:
                ponencia_sala=Ponencia.objects.filter(sala=sala).first()
            if ponencia_sala:
                salas_env.append({
                    'sala':sala,
                    'ponencia':ponencia_sala,
                    'ponentes':Ponente.objects.filter(ponencia_ponente__pk=ponencia_sala.id).distinct() ,
                        })
        context['salas']=salas_env      
        if congreso is not None:
           
            context['congreso']=congreso
            #context['dias_faltan']=date.today()-self.model.fecha_inicio
            with connections['default'].cursor() as cursor:
                    sql_query = '''SELECT DISTINCT fecha_inicio::date FROM public."MedCongressApp_ponencia" where published is TRUE  and congreso_id= '''+ str(congreso.id) +''' ORDER by fecha_inicio'''
                    cursor.execute(sql_query)
                    data2 = [row[0] for row in cursor.fetchall()]
            with connections['default'].cursor() as cursor:
                    sql_query = '''SELECT DISTINCT fecha_inicio::date FROM public."MedCongressApp_taller" where published is TRUE  and congreso_id= '''+ str(congreso.id) +''' ORDER by fecha_inicio'''
                    cursor.execute(sql_query)
                    data1 = [row[0] for row in cursor.fetchall()]
                    
            
            data=data2+[i for i in data1 if i not in data2]
           
            context['fecha_ponencias']= data
            ponencias_env=[] 
           
            for dat in context['fecha_ponencias'] :
                bloques=Bloque.objects.filter(fecha_inicio__date=dat,congreso=congreso,published=True).order_by('fecha_inicio')
                ponencias=Ponencia.objects.filter(fecha_inicio__date=dat,congreso=congreso,published=True,bloque=None).order_by('fecha_inicio')
                talleres=Taller.objects.filter(fecha_inicio__date=dat,congreso=congreso,published=True,bloque=None).order_by('fecha_inicio')
                result=[]
                for bloque in bloques: 
                    bloque_ponencias=Ponencia.objects.filter(bloque=bloque,published=True).order_by('fecha_inicio')
                   
                    bloque_talleres=Taller.objects.filter(bloque=bloque,published=True).order_by('fecha_inicio')
                    eventos=[]
                    for ponencia in bloque_ponencias: 
                        eventos.append({
                        'id':ponencia.id,
                        'path':ponencia.path,
                        'titulo': ponencia.titulo,
                        'ver_ponencia':ponencia.cod_video,
                        'fecha_inicio': ponencia.fecha_inicio ,# una relación a otro modelo
                        'detalle':ponencia.detalle ,
                        'ponentes':Ponente.objects.filter(ponencia_ponente__pk=ponencia.id).distinct() ,
                        'tipo':'Ponencia',
                        'sala':ponencia.sala,
                        })
                    for taller in bloque_talleres: 
                        eventos.append({
                        'id':taller.id,
                        'path':taller.path,
                        'titulo': taller.titulo,
                        'ver_ponencia':taller.cod_video,
                        'fecha_inicio': taller.fecha_inicio ,# una relación a otro modelo
                        'detalle':taller.detalle ,
                        'ponentes':Ponente.objects.filter(taller_ponente__pk=taller.id).distinct() ,
                        'tipo':'Taller',# la misma relación, otro campo
                        })
                    eventos = sorted(eventos, key=lambda k: k['fecha_inicio'])
                    result.append({
                    'id':bloque.id,
                    'path':bloque.path,
                    'cod_video':bloque.cod_video,
                    'moderador':Moderador.objects.filter(bloque_moderador__pk=bloque.id).distinct() ,
                    'titulo': bloque.titulo,
                    'fecha_inicio': bloque.fecha_inicio ,# una relación a otro modelo
                    'detalle':bloque.detalle ,
                    'eventos':eventos,
                    'tipo':'Bloque',# la misma relación, otro campo
                    })
                   
                for ponencia in ponencias: 
                    result.append({
                    'id':ponencia.id,
                    'path':ponencia.path,
                    'ver_ponencia':ponencia.cod_video,
                    'titulo': ponencia.titulo,
                    'fecha_inicio': ponencia.fecha_inicio ,# una relación a otro modelo
                    'detalle':ponencia.detalle ,
                    'ponentes':Ponente.objects.filter(ponencia_ponente__pk=ponencia.id).distinct() ,
                    'tipo':'Ponencia',# la misma relación, otro campo
                    'sala':ponencia.sala,
                    })
                for taller in talleres: 
                    result.append({
                    'id':taller.id,
                    'path':taller.path,
                    'titulo': taller.titulo,
                    'ver_ponencia':taller.cod_video,
                    'fecha_inicio': taller.fecha_inicio ,# una relación a otro modelo
                    'detalle':taller.detalle ,
                    'ponentes':Ponente.objects.filter(taller_ponente__pk=taller.id).distinct() ,
                    'tipo':'Taller',# la misma relación, otro campo
                    })
                result = sorted(result, key=lambda k: k['fecha_inicio'])
              
                # ponentes_env.append(Ponente.objects.filter(ponencia_ponente__pk=ponencia.id).distinct()) 
                ponencias_env.append(result)
                # for taller in talleres:
                #     ponentes_env.append(Taller.objects.filter(reltallerponente__pk=taller.id).distinct()) 
             
                #     ponencias_env.append(talleres)
          
            context['ponencias']=ponencias_env
    
            prueba_ponecia=Ponencia.objects.filter(congreso=congreso.pk,published=True)
            id_p=[]
            for pp in prueba_ponecia:
                id_p.append(pp.pk)

            ponentes=Ponente.objects.filter(ponencia_ponente__in=id_p).distinct()
            ponentes_env=[]
            for ponente in ponentes: 
                ponentes_env.append({
                'id':ponente.id,
                'id_user':ponente.user.pk,
                'nombre': ponente.user.usuario.first_name,
                'apellido': ponente.user.usuario.last_name ,
                'foto':ponente.user.foto,
                'tipo':'Ponente',
                })
            ponencias_video_env=[]

            fechas_ponencias_video=Ponencia.objects.filter(congreso=congreso.pk,published=True).exclude(cod_video='').distinct('fecha_inicio__date').values('fecha_inicio__date')
            fechas_talleres_video=Taller.objects.filter(congreso=congreso.pk,published=True).exclude(cod_video='').distinct('fecha_inicio__date').values('fecha_inicio__date')
            
            fechas=[]
            for fecha_ponencia in fechas_ponencias_video:
                fechas.append(fecha_ponencia['fecha_inicio__date'])
            for fechas_tallere in fechas_talleres_video:
                fechas.append(fechas_tallere['fecha_inicio__date'])
            
            fechas=set(fechas)
            fechas_final=sorted(fechas)
            
            for j in range(0, len(fechas_final)):
                ponencias_video= Ponencia.objects.filter(congreso=congreso.pk,published=True,fecha_inicio__date=fechas_final[j]).exclude(cod_video='')
                talleres_video= Taller.objects.filter(congreso=congreso.pk,published=True,fecha_inicio__date=fechas_final[j]).exclude(cod_video='')
                ponencias_video_env.append({'fecha':fechas_final[j],
                                            'ponencias':ponencias_video,
                                            'talleres':talleres_video})
            context['ponencias_video']=ponencias_video_env
            prueba_taller=Taller.objects.filter(congreso=congreso.pk,published=True)
            id_t=[]
            for pp in prueba_taller:
                id_t.append(pp.pk)

            ponentes=Ponente.objects.filter(taller_ponente__in=id_t).distinct()

            for ponente in ponentes: 
                var=False
                for pon in ponentes_env:
                    if pon['id_user']==ponente.user.pk:
                        var=True
                if not var:
                    ponentes_env.append({
                    'id':ponente.id,
                    'id_user':ponente.user.pk,
                    'nombre': ponente.user.usuario.first_name,
                    'apellido': ponente.user.usuario.last_name ,
                    'foto':ponente.user.foto,
                    'tipo':'Ponente',
                    })
            
            bloques=Bloque.objects.filter(congreso=congreso.pk,published=True)
            id_b=[]
            for pp in bloques:
              
                id_b.append(pp.pk)
            moderadores=Moderador.objects.filter(bloque_moderador__in=id_b).distinct()

            for moderador in moderadores: 
                var=False
                for pon in ponentes_env:
                    if pon['id_user']==moderador.user.pk:
                        var=True
                if not var:
                    ponentes_env.append({
                    'id':moderador.id,
                    'id_user':moderador.user.pk,
                    'nombre': moderador.user.usuario.first_name,
                    'apellido': moderador.user.usuario.last_name ,
                    'foto':moderador.user.foto,
                    'tipo':'Moderador',
                    })


            context['ponentes_congreso']=ponentes_env
            cat_pago=RelCongresoCategoriaPago.objects.filter(congreso=congreso.pk)
            context['cat_ponente']=RelPonenciaPonente.objects.all()

            if self.request.user.is_authenticated :
                user_perfil=PerfilUsuario.objects.filter(usuario=self.request.user.pk).first()
                context['cuestionario_aprobado']=RelCongresoUser.objects.filter(user=user_perfil.pk, congreso=congreso.pk,is_constancia=True).exists()
                talleres=Taller.objects.filter(congreso=congreso.pk,published=True).order_by('fecha_inicio')
                ver=[]
                for taller in talleres:
                    if RelTalleresCategoriaPago.objects.filter(taller=taller).exists():
                        cat_pa=RelTalleresCategoriaPago.objects.filter(taller=taller)
                    else:
                        cat_pa=True
                    
                    if RelTallerUser.objects.filter(user=user_perfil.pk, taller=taller.pk).exists():
                        ver.append([taller,cat_pa,True])
                    else:
                        ver.append([taller,cat_pa,False])  
                context['talleres']=ver
                pagos = RelCongresoUser.objects.filter(user=user_perfil.pk, congreso=congreso.pk).order_by('precio')
                
                if pagos.exists():
                    pagos_p = RelCongresoUser.objects.filter(user=user_perfil.pk, congreso=congreso.pk,is_pagado=True).order_by('precio') 
                    if pagos_p.exists():
                        context['permiso'] = True
                    else:
                        context['permiso'] = True
                else: 
                    context['permiso'] = False                                                                  
            else:
                talleres=Taller.objects.filter(congreso=congreso.pk,published=True).order_by('fecha_inicio')
                ver=[]
                for taller in talleres:
                    if RelTalleresCategoriaPago.objects.filter(taller=taller).exists():
                        cat_pa=RelTalleresCategoriaPago.objects.filter(taller=taller)
                    else:
                        cat_pa=True
                    ver.append([taller,cat_pa,False])      
                context['talleres']=ver
                context['permiso'] = False  
            context['categorias_pago']=cat_pago
            context['programas']=DocumentoPrograma.objects.filter(congreso=congreso)
            context['trabajos']=TrabajosInvestigacion.objects.filter(congreso=congreso)
            context['preg_frecuentes']=PreguntasFrecuentes.objects.filter(congreso=congreso,published=True)
            context['cuestionario']=CuestionarioPregunta.objects.filter(congreso=congreso,published=True).exists()
           
        return context

class AsignarConstanciasUsuario(validarOrganizador,FormView):
    
    form_class = AsignarConstanciaUserForms
    template_name = 'MedCongressAdmin/asig_constancia_usuario.html'

    def get_success_url(self):
        url =  reverse_lazy('MedCongressAdmin:asig_congress_list' )
        self.success_url='%s?search=%s'%(url,self.request.GET.get('search'))
        return self.success_url

    def form_valid(self, form):
        relacion=RelCongresoUser.objects.get(pk=self.kwargs.get('pk'))
        relacion.foto_constancia=self.request.FILES['foto_constancia']
        relacion.is_constancia=True
        relacion.fecha_constancia=datetime.now()
        score=0
        if relacion.congreso.score:
            score=relacion.congreso.score
        if relacion.user.score is None:
            relacion.user.score=score
        else:
            relacion.user.score= relacion.user.score+score
        relacion.user.save()
                                       
        relacion.save()
        
        email = EmailMessage('Constancia', '''Estimado usuario de MedCongress. Por medio del presente, se le informa que la constancia de asistencia al congreso  '''+relacion.congreso.titulo +''', el cual esperamos haya sido de tu interés y agrado, ya está en su perfil en nuestra plataforma https://medcongress.com.mx/ 
        
        Recuerda que tienes acceso a las presentaciones del simposio a través de la plataforma de MedCongress, solo tienes que ingresar a la página https://medcongress.com.mx/ y en login ingresar tu correo electrónico y tu contraseña, dentro del programa podrás elegir las ponencias de tu interés que deseas ver.''', to = [relacion.user.usuario.email] )

        email.send()

        return super(AsignarConstanciasUsuario, self).form_valid(form)
   

    # def get_success_url(self):
    #     if self.kwargs.get('pk'):
    #         pregunta=CuestionarioPregunta.objects.get(pk=self.kwargs.get('pk'))
            
    #         self.success_url =  reverse_lazy('MedCongressAdmin:Congres_cuestionario',kwargs={'path': pregunta.congreso.path} )
    #     return self.success_url 

class AsignarConstancias(validarOrganizador,FormView):
    template_name = 'MedCongressAdmin/asig_constancia.html'
    form_class = AsignarConstanciasForm
    success_url = reverse_lazy('MedCongressAdmin:asig_constancia_list')

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs(*args, **kwargs)
        if self.request.user.is_staff:
            kwargs['user'] = False
        else:
            kwargs['user'] =self.request.user
        return kwargs


    def form_valid(self, form):
        titulo= self.request.POST['congresos']
        t_user= int(self.request.POST['tipo_usuario'])
        n_user='' 
        folio_ini=None
        folio_fin=None
        folio_dis=None
        congreso=Congreso.objects.get(pk=self.request.POST['congresos'])
        if self.request.POST.get('folio'):
            folio_ini=int(self.request.POST['folio_ini'])
            folio_fin=int(self.request.POST['folio_fin'])
            folio_dis=self.request.POST['folio_dis']

        prueba=Constancia.apply_async(args=[titulo,t_user,folio_ini,folio_fin,folio_dis])
        respuesta=prueba.get()
        if respuesta['success'] :
            messages.success(self.request,respuesta['mensaje'])
        else:
            messages.warning(self.request,respuesta['mensaje']) 
        return super().form_valid(form)

    # def get_context_data(self, **kwargs):
    #     context = super(AsignarConstancias, self).get_context_data(**kwargs)
    #     if self.request.user.is_staff:
    #         congreso=Congreso.objects.all()
    #     else:
    #         congreso=[]
    #         congresos=Organizador.objects.filter(user=self.request.user.perfilusuario)
    #         for cong in congresos:
    #             congreso.append(cong.congreso)
    #     context['congresos']=congreso
    #     return context
         

    # def post(self, request, **kwargs):
    #     titulo= self.request.POST['my_congress']
    #     t_user= int(self.request.POST['tipo_usuario'])
    #     n_user='' 
    #     folio_ini=None
    #     folio_fin=None
    #     folio_dis=None
    #     congreso=Congreso.objects.get(pk=self.request.POST['my_congress'])
    #     if congreso: 
    #         if self.request.POST.get('folio'):
    #             folio_ini=int(self.request.POST['folio_ini'])
    #             folio_fin=int(self.request.POST['folio_fin'])
    #             folio_dis=self.request.POST['folio_dis']
    #             for num_folio in range(folio_ini,folio_fin):
    #                 print('folio_completo=%s '%(folio_dis.replace('#',str(num_folio))))
    #                 if RelCongresoUser.objects.filter(congreso=congreso,folio_constancia=folio_dis.replace('#',str(num_folio))).exists() or ConstanciaUsuario.objects.filter(congreso=congreso,folio_constancia=folio_dis.replace('#',str(num_folio))).exists():
    #                     messages.warning(self.request,'El folio %s ya esta asignado en este congreso'%(folio_dis.replace('#',str(num_folio))))
    #                     return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list'))
                        
    #         if t_user==1:
    #             n_user='Participantes'
    #             if not congreso.foto_constancia:                                 
    #                 messages.warning(self.request,'Error.....Ese Congreso no tiene asignada ninguna foto para la constancia del Participante')
    #                 return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list'))
    #             if self.request.POST.get('folio'):

    #         elif t_user==2:
    #             n_user='Ponentes'
    #             if not congreso.foto_const_ponente:                                 
    #                 messages.warning(self.request,'Error.....Ese Congreso no tiene asignada ninguna foto para la constancia del Ponente')
    #                 return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list'))
    #         elif t_user==3:
    #             n_user='Moderadores'
    #             if not congreso.foto_const_moderador:                                 
    #                 messages.warning(self.request,'Error.....Ese Congreso no tiene asignada ninguna foto para la constancia del Moderador')
    #                 return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list'))
    #     else:
    #         messages.warning(self.request,'Ese Congreso no existe')
    #         return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list')) 
    #     print('envio al celery')
    #     prueba=Constancia.apply_async(args=[titulo,t_user,folio_ini,folio_fin,folio_dis])
    #     respuesta=prueba.get()
    #     if respuesta['success'] :
    #         messages.success(self.request,respuesta['mensaje'])
    #     else:
    #         messages.warning(self.request,respuesta['mensaje']) 
    #     return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_list'))

class vTableAsJSONAsigCongreso(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['user__usuario__first_name','user__usuario__email','congreso__titulo','cantidad','categoria_pago__nombre','is_pagado','is_constancia']
           
        #listado que muestra en dependencia de donde estes parado
        if request.user.is_staff:
            object_list = RelCongresoUser.objects.filter(is_pagado=True)
        else:
            congreso_orgs=Organizador.objects.filter(user=request.user.perfilusuario)
            object_list=RelCongresoUser.objects.filter(pk=0)
            for organizador in congreso_orgs:
                object_list |=RelCongresoUser.objects.filter(congreso=organizador.congreso,is_pagado=True)

        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        
        sort_colr = col_name_map[sort_col]
        object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(user__usuario__last_name__icontains=search_text) | Q(user__usuario__email__icontains=search_text)|Q(user__usuario__first_name__icontains=search_text)|Q(congreso__titulo__icontains=search_text)|Q(cantidad__icontains=search_text)|Q(categoria_pago__nombre__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
       
            # if objet.ponente:
            #     user= '%s %s'%(objet.ponente.first().user.usuario.first_name,objet.ponente.first().user.usuario.last_name)
           
           #Guardar datos en un dic 
        for objet in filtered_object_list[start:(start+delta)]:
           
           
            constancia='Si'
            
            if objet.foto_constancia or RelCongresoUser.objects.filter(congreso=objet.congreso,user=objet.user,is_constancia=True).exists():
                        constancia='''Si'''                                         
            else:
                constancia= '''<a href="'''+ reverse('MedCongressAdmin:constancia_usuario_add',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''"
                                                    title="Asignar Constancia">
                                                    <i class="icon icon-constancia"></i>
                                                </a>'''   
            cat_pago='Beca'
            if objet.categoria_pago:
                cat_pago=objet.categoria_pago.nombre
            enviar.append({ 'usuario':'%s %s'%(objet.user.usuario.first_name,objet.user.usuario.last_name),
                            'email':'<p class="text"  >'+ objet.user.usuario.email+'</p>',
                            'congreso' : objet.congreso.titulo,
                            'cantidad' : objet.cantidad,
                            'cat_pago':cat_pago,
                            'constancia':constancia,
                            'operaciones' : 
                                                    '''<a id="del_'''+ str(objet.pk)+'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk)+''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>''',
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)    

class vTableAsJSONCongresos(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['titulo','published']
           
        #listado que muestra en dependencia de donde estes parado
        if request.user.is_staff:
            object_list = Congreso.objects.all()
        else:
            congreso_orgs=Organizador.objects.filter(user=request.user.perfilusuario)
            object_list=Congreso.objects.filter(pk=0)
            for organizador in congreso_orgs:
                object_list |=Congreso.objects.filter(pk=organizador.congreso.pk)
        
        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        
        sort_colr = col_name_map[sort_col]
        object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(titulo__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
       
            # if objet.ponente:
            #     user= '%s %s'%(objet.ponente.first().user.usuario.first_name,objet.ponente.first().user.usuario.last_name)
           
           #Guardar datos en un dic 
        for objet in filtered_object_list[start:(start+delta)]:
            pagado='No'
            if objet.published:
                pagado='Si'
            operaciones=''' <a 
                                                    href="javascript:editItem('''+ str(objet.pk)+''')"
                                                        title="Editar" style="margin-left: 5px;"><i class="icon-editar" style="padding: 15px;"></i></a>'''
            if request.user.is_staff:
                operaciones+= '''<a id="del_'''+ str(objet.pk)+'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk)+''')"
                                                        title="Eliminar" style="margin-left: 5px;">
                                                        <i class="icon-eliminar" style="padding: 15px;"></i>
                                                    </a>'''
            enviar.append({ 'titulo':objet.titulo,
                            'publicado': pagado,
                           
                            'programa' : ''' <a id=""
                                                    href="'''+ reverse('MedCongressAdmin:bloques_list')+'''?congreso='''+ str(objet.path)+'''"
                                                    title="Bloques " style="margin-left: 5px;">
                                                    <i class="icon-block " style=" padding:15px ;"> </i>
                                                </a>
                                                
                                                <a id=""
                                                    href="'''+ reverse('MedCongressAdmin:ponencias_list')+'''?congreso='''+ str(objet.path)+'''"
                                                    title="Ponencias " style="margin-left: 5px;">
                                                    <i class="icon-ponencia " style=" padding:15px"> </i>
                                                </a>
                                               
                                                <a id=""
                                                     href=" '''+ reverse('MedCongressAdmin:talleres_list')+'''?congreso='''+ str(objet.path)+'''"
                                                    title="Talleres " style="margin-left: 5px;">
                                                    <i class="icon-taller " style=" padding:15px"> </i>
                                                </a>
                                                
                                                <a id=""
                                                href="'''+ reverse('MedCongressAdmin:Congres_cuestionario',kwargs={'path':objet.path})+'''"
                                               title="Cuestionario " style="margin-left: 5px;">
                                               <i class="icon-cuestionario " style=" padding:15px"> </i>
                                           </a>
                                           
                                           <a id=""
                                           href="'''+ reverse('MedCongressAdmin:Congres_programas',kwargs={'path':objet.path})+'''"
                                          title="Documentos de Programa " style="margin-left: 5px;">
                                          <i  class="icon icon-docu" > </i>
                                      </a>
                                       <a id=""
                                           href="'''+ reverse('MedCongressAdmin:Congres_salas',kwargs={'path':objet.path})+'''"
                                          title="Salas " style="margin-left: 5px;">
                                          <i  class="icon icon-sala" > </i>
                                      </a>''',
                                      
                            'otros'     : '''
                                                    <a id=""
                                                    href=" '''+ reverse('MedCongressAdmin:Congres_trabajos',kwargs={'path':objet.path})+'''"
                                                        title="Trabajos Investigativos" style="margin-left: 5px;">
                                                        <i class=" icon icon-inv " > </i>
                                                    </a>  
                                                     
                                                    <a id=""
                                                    href="'''+ reverse('MedCongressAdmin:Congres_freg_frecuente',kwargs={'path':objet.path})+'''"
                                                        title="Preguntas Frecuentes" style="margin-left: 5px;">
                                                        <i class="icon-preguntas" style=" padding:15px"> </i>
                                                    </a> 
                                                   
                                                    <a id=""
                                                    href=" '''+ reverse('MedCongressAdmin:Congres_pagos',kwargs={'path':objet.path})+'''"
                                                        title="Categorias de pago" style="margin-left: 5px;">
                                                        <i class="icon-pago" style=" padding:15px"> </i>
                                                    </a> 
                                                     
                                                    <a id=""
                                                    href="'''+ reverse('MedCongressAdmin:patrocinadores_list')+'''?congreso='''+ str(objet.path)+'''"
                                                        title="Patrocinadores" style="margin-left: 5px;">
                                                        <i class="icon-patrocinador" style=" padding:15px"> </i>
                                                    </a>
                                                   
                                                    <a id=""
                                                    href=" '''+ reverse('MedCongressAdmin:socios_list')+'''?congreso='''+ str(objet.path)+'''"
                                                        title="Socios" style="margin-left: 5px;">
                                                        <i class="icon-socio" style=" padding:15px"> </i>
                                                    </a>
                                                    
                                                    <a id="" target="_blank"
                                                    href=" '''+ reverse('MedCongressAdmin:congress_previsualizar',kwargs={'path':objet.path})+'''"
                                                        title="Previsualizar" style="margin-left: 5px;">
                                                        <i class="icon icon-visualizar" > </i>
                                                    </a> ''',
                           
                            'operaciones' : operaciones,
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)    

class LogsCongreso(validarUser,FormView):
    form_class=ExportarLogsCongresoExelForm
    template_name = 'MedCongressAdmin/congress/log_form.html'
    def form_valid(self, form):
        id_congreso=self.request.POST['congreso']
        congreso= Congreso.objects.get(pk=id_congreso)
        if self.request.POST['fecha_fin']:
            fecha_fin=self.request.POST['fecha_fin']
            d_fecha_final = datetime.strptime(fecha_fin,'%Y-%m-%d') + timedelta(days=1)
        else:
             d_fecha_final = datetime.now() + timedelta(days=1)   
        fecha_inicio=self.request.POST['fecha_inicio']
        if fecha_inicio:
            d_fecha_inicio = datetime.strptime(fecha_inicio,'%Y-%m-%d')
          
            query=UserActivityLog.objects.filter(congreso=id_congreso,fecha__range=[d_fecha_inicio,d_fecha_final]).order_by('user','fecha')
        else:
            
            query=UserActivityLog.objects.filter(congreso=id_congreso,fecha__lt=d_fecha_final).order_by('user','fecha')
        
        if query:
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=25
            ws.column_dimensions['C'].width=20
            ws.column_dimensions['D'].width=150
            

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            celdas_fecha = NamedStyle(name="celdas_fecha")
            celdas_fecha.font=Font(size=12)
            
            celdas_fecha.alignment=Alignment(horizontal='center',mergeCell=True)
            celdas_fecha.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            label = NamedStyle(name="label")
            label.font=Font(size=12,bold=True)
            label.alignment=Alignment(horizontal='right',mergeCell=True)
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Logs del Congresos :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            ws['A2'] ='" %s "'%(congreso.titulo) 
            ws['A2'].font = Font(size=12,bold=True)
            ws['A2'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            cont=3
            num=1
            id_user=0
            if query:
                for quer in query:
                    if id_user!=quer.user.pk:
                        cont = cont + 1
                        num=1
                        ws.merge_cells('A%s:B%s'%(cont,cont))
                        ws.merge_cells('C%s:D%s'%(cont,cont))
                        ws.cell(row=cont,column=1).style=label
                        ws['A%s'%(cont)] = 'Usuario:'
                        ws['C%s'%(cont)] = '%s %s <<%s>>'%(quer.user.usuario.first_name,quer.user.usuario.last_name,quer.user.usuario.email)
                        cont = cont + 1
                        ws.merge_cells('A%s:B%s'%(cont,cont))
                        ws.merge_cells('C%s:D%s'%(cont,cont))
                        ws.cell(row=cont,column=1).style=label
                        ws['A%s'%(cont)] = 'Categoría:'
                        ws['C%s'%(cont)] = '%s'%(quer.user.categoria)
                        cont = cont + 1
                        
                        ws.cell(row=cont,column=1).style=titulo
                        ws.cell(row=cont,column=2).style=titulo
                        ws.cell(row=cont,column=3).style=titulo
                        ws.cell(row=cont,column=4).style=titulo
                        
                        ws.cell(row=cont,column=1).value='No.'
                        ws.cell(row=cont,column=2).value='Fecha' 
                        ws.cell(row=cont,column=3).value='Tiempo (H:M:S)'
                        ws.cell(row=cont,column=4).value='Acción'
                        id_user= quer.user.pk   
                    
                        cont = cont + 1
                #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
                
                    
                    ws.cell(row=cont,column=1).style=celdas_fecha
                    ws.cell(row=cont,column=1).value = num
                    ws.cell(row=cont,column=2).style=celdas_fecha
                    ws.cell(row=cont,column=2).value = quer.fecha
                    ws.cell(row=cont,column=3).style=celdas_fecha
                    tiempo=quer.tiempo.split('.')
                    ws.cell(row=cont,column=3).value =tiempo[0]
                    ws.cell(row=cont,column=4).style=celdas
                    ws.cell(row=cont,column=4).value = '  %s'%(quer.mensaje) 
                    
                    cont = cont + 1
                    num = num + 1
            else:
                ws.merge_cells('A%s:F%s'%(cont,cont))
                ws.cell(row=cont,column=1).value = 'Este congreso no tiene Logs'
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=LogsCongreso.xlsx"
            wb.save(response)
            return response
        
        else:
            messages.warning(self.request, 'No existen Logs con estos criterios de busquedas')
            return HttpResponseRedirect(reverse('MedCongressAdmin:LogsCongreso'))
  

