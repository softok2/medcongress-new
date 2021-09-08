import json
import base64 
from os import remove
from pathlib import Path
from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin
from django.db.models import Sum
from django.http import (HttpResponse, HttpResponseBadRequest,
                         HttpResponseRedirect)
from django.urls import reverse, reverse_lazy
from django.utils.crypto import get_random_string
from django.views.generic import FormView, ListView, TemplateView
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from MedCongressAdmin.forms.congres_forms import (AsignarTallerForms,
                                                  ExportarTallerExelForm,
                                                  PonenteTallerForm,
                                                  TallerCategPagoForm,
                                                  TallerForms)
from MedCongressApp.models import (Bloque, Congreso, Ponente,Organizador,
                                   RelTalleresCategoriaPago, RelTallerPonente,
                                   RelTallerUser, Taller, Ubicacion)
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side,NamedStyle)
from MedCongressAdmin.apps import validarUser,validarOrganizador
from MedCongressAdmin.task import Constanciataller
from django.db.models import Q


class TalleresListView(validarOrganizador,TemplateView):
    model = Taller
    context_object_name = 'talleres'
    template_name = 'MedCongressAdmin/taller/listar.html'
    def get(self, request, **kwargs):
        if self.request.GET.get('congreso'):
            congreso=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            if congreso is None:
                return   HttpResponseRedirect(reverse('Error404'))
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff: 
                return   HttpResponseRedirect(reverse('Error403'))
        elif self.request.GET.get('bloque'):
            bloque=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            if bloque is None:
                return   HttpResponseRedirect(reverse('Error404'))
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=bloque.congreso).exists() and not self.request.user.is_staff: 
                return   HttpResponseRedirect(reverse('Error403')) 
        else:
            if not self.request.user.is_staff:
                return   HttpResponseRedirect(reverse('Error403'))   
        return self.render_to_response(self.get_context_data())
    def get_context_data(self, **kwargs):
        context=super(TalleresListView,self).get_context_data(**kwargs)
        context['search']=self.request.GET.get('search')
        context['talleres']=Taller.objects.all()
        if self.request.GET.get('congreso'):
            context['congreso']=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            context['talleres']=Taller.objects.filter(congreso= context['congreso'])
        if self.request.GET.get('bloque'):
            context['bloque']=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            context['talleres']=Taller.objects.filter(bloque= context['bloque'])
            if self.request.GET.get('congreso_bloque'):
                context['congreso_bloque']=True
        return context

class  TallerCreateView(validarOrganizador,FormView):
    form_class = TallerForms
    success_url = reverse_lazy('MedCongressAdmin:talleres_list')
    template_name = 'MedCongressAdmin/taller/form.html'

    def get(self, request, **kwargs):

        if self.request.GET.get('congreso'):
            congreso=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            if congreso is None:
                return   HttpResponseRedirect(reverse('Error404'))
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff:
                return   HttpResponseRedirect(reverse('Error403'))
        
        elif self.request.GET.get('bloque'):
            bloque=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=bloque.congreso).exists() and not self.request.user.is_staff: 
                return   HttpResponseRedirect(reverse('Error403'))
            if bloque is None:
                return   HttpResponseRedirect(reverse('Error404'))
        else:
            if not self.request.user.is_staff:
                return   HttpResponseRedirect(reverse('Error403'))
        return self.render_to_response(self.get_context_data())

    def form_valid(self, form):
        if not self.request.POST.getlist('taller_ponente-ponente'):
            messages.warning(self.request, 'Debe al menos entrar un ponente')
            return super().form_invalid(form) 
        taller=form['taller'].save(commit=False)
        ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)

        if ubic.exists():
            taller.lugar=ubic.first()
        else:
            ubicacion=form['ubicacion'].save(commit=True)
            taller.lugar=ubicacion
        path=taller.titulo.replace("/","").replace(" ","-").replace("?","").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ñ","n")
        chars = '0123456789'
        secret_key = get_random_string(5, chars)
        taller.path=path+secret_key
        id_video=['']
        if taller.cod_video:
            id_video=taller.cod_video.split(sep='https://player.vimeo.com/video/')
            id_video=id_video[-1].split(sep='"')
        taller.id_video=id_video[0]
        image_64_encode=self.request.POST['taller-prueba']
        campo = image_64_encode.split(",")
        image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
        chars = '0123456789'
        nombre = get_random_string(5, chars)
        image_result = open('MedCongressApp/static/talleres/imagen_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
        image_result.write(image_64_decode)
        taller.imagen='talleres/imagen_%s.png'%(nombre)
        if self.request.POST['taller-constancia']:
                image_64_encode=self.request.POST['taller-constancia']
                campo = image_64_encode.split(",")
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_result = open('MedCongressApp/static/congreso/img_constancia/imagen_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                taller.foto_constancia='congreso/img_constancia/imagen_%s.png'%(nombre)
        taller.save()
        for ponente in self.request.POST.getlist('taller_ponente-ponente'):
                ponente_=Ponente.objects.get(pk=ponente)
                po= RelTallerPonente(ponente=ponente_,taller=taller)
                po.save()
        return super(TallerCreateView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        
        context = super(TallerCreateView, self).get_context_data(**kwargs)
        if self.request.GET.get('congreso'):
            context['con']=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            context['blo']=Bloque.objects.filter(congreso=context['con'])
        if self.request.GET.get('bloque'):
            context['bloque']=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            context['congreso']=context['bloque'].congreso
            context['blo']= None
        if self.request.GET.get('congreso_bloque'):
            context['congreso_bloque']=True
        return context 
    def get_success_url(self):
        url=reverse_lazy('MedCongressAdmin:talleres_list')
        self.success_url='%s?&search=%s'%(url,self.request.GET.get('search'))
        if self.request.GET.get('congreso'):
            self.success_url =  '%s?congreso=%s&search=%s'%(url,self.request.GET.get('congreso'),self.request.GET.get('search')) 
        if self.request.GET.get('bloque'): 
            self.success_url =  '%s?bloque=%s&search=%s'%(url,self.request.GET.get('bloque'),self.request.GET.get('search')) 
            if self.request.GET.get('congreso_bloque'):
                self.success_url =  '%s?bloque=%s&search=%s&congreso_bloque=true'%(url,self.request.GET.get('bloque'),self.request.GET.get('search')) 

        return self.success_url 

########## Vista de las Categorias de Pago de un Congreso #############

class TallerCategPagosListView(validarOrganizador,TemplateView):
    template_name= 'MedCongressAdmin/taller/listar_cat_pagos.html' 
    

    def get(self, request, **kwargs):
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        if taller is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=taller.congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        return self.render_to_response(self.get_context_data())    
    def get_context_data(self, **kwargs):
        context = super(TallerCategPagosListView, self).get_context_data(**kwargs)
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        context['congres']=taller
        context['cat_pagos']=RelTalleresCategoriaPago.objects.filter(taller=taller)
        if self.request.GET.get('congreso'):
            context['congreso']=taller.congreso   
        if self.request.GET.get('bloque'):
            context['bloque']=taller.bloque
        if self.request.GET.get('congreso_bloque'):
            context['congreso']=taller.congreso
            context['bloque']=taller.bloque
            context['congreso_bloque']=True
        return context        
        
class  TallerCategPagosCreateView(validarOrganizador,CreateView):
    info_sended =Taller()
    form_class = TallerCategPagoForm
    # success_url = reverse_lazy('MedCongressAdmin:ponencias_list')
    template_name = 'MedCongressAdmin/taller/cat_pago_form.html'

    def get(self, request, **kwargs):
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        if taller is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=taller.congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        self.object=taller
        return self.render_to_response(self.get_context_data()) 

    def form_valid(self, form):
        congreso=form.save(commit=False)
  
        congreso.save()
        return super(TallerCategPagosCreateView, self).form_valid(form)

    def get_success_url(self):
        url =  reverse_lazy('MedCongressAdmin:Taller_pagos',kwargs={'path': self.kwargs.get('path')} )+'?&search=%s'%(self.request.GET.get('search'))
        if self.request.GET.get('congreso'):
            self.success_url =  '%s&congreso=True'%(url) 
        if self.request.GET.get('bloque'): 
            self.success_url =  '%s&bloque=True'%(url) 
        if self.request.GET.get('congreso_bloque'):
            self.success_url =  '%s&congreso_bloque=True'%(url) 
        return self.success_url

    def get_context_data(self, **kwargs):
        ctx = super(TallerCategPagosCreateView, self).get_context_data(**kwargs)
        pon=Taller.objects.filter(path=self.kwargs.get('path')).first()
        ctx['taller'] = pon
        if self.request.GET.get('congreso'):
            ctx['congreso']=pon.congreso   
        if self.request.GET.get('bloque'):
            ctx['bloque']=pon.bloque
        if self.request.GET.get('congreso_bloque'):
            ctx['congreso']=pon.congreso
            ctx['bloque']=pon.bloque
            ctx['congreso_bloque']=True
                
        return ctx

class TallerUpdateView(validarOrganizador,FormView):
    form_class = TallerForms
    success_url = reverse_lazy('MedCongressAdmin:talleres_list')
    template_name = 'MedCongressAdmin/taller/form.html'
    def get(self, request, **kwargs):

        if self.request.GET.get('congreso'):
            congreso=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            if congreso is None:
                return   HttpResponseRedirect(reverse('Error404'))
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=congreso).exists() and not self.request.user.is_staff:
                return   HttpResponseRedirect(reverse('Error403'))
        
        elif self.request.GET.get('bloque'):
            bloque=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=bloque.congreso).exists() and not self.request.user.is_staff: 
                return   HttpResponseRedirect(reverse('Error403'))
            if bloque is None:
                return   HttpResponseRedirect(reverse('Error404'))
        else:
            if not self.request.user.is_staff:
                return   HttpResponseRedirect(reverse('Error403'))
        return self.render_to_response(self.get_context_data())
    def get_queryset(self, **kwargs):
        return Taller.objects.filter(pk=self.kwargs.get('pk'))
    
    def get_form_kwargs(self):
        taller=Taller.objects.get(pk=self.kwargs.get('pk'))
        self.object=taller
        kwargs = super(TallerUpdateView, self).get_form_kwargs()
        kwargs.update(instance={
            'taller': self.object,
            'ubicacion': self.object.lugar,
        })
        return kwargs

    def get_context_data(self, **kwargs):
        taller=Taller.objects.get(pk=self.kwargs.get('pk'))
        self.object=taller
        context=super().get_context_data(**kwargs)
        if self.object.meta_og_imagen:
            context['imagen_meta']='/static/%s'%(self.object.meta_og_imagen)
        if self.object.imagen:
            context['imagen_seg_url']=self.object.imagen
        context['bloque_update']=self.object.bloque
        if self.object.foto_constancia:
            context['foto_constancia']='/static/%s'%(self.object.foto_constancia)
        context['update']='update'
        ponentes=Ponente.objects.all()
        relaciones=RelTallerPonente.objects.filter(taller=taller)
        ponentes_env=[]
        activo=False
        for ponente in ponentes:
            activo=False
            for relacion in relaciones:
                if relacion.ponente.pk==ponente.id:
                    activo=True
                
            ponentes_env.append({'id':ponente.id,
            'nombre':'%s %s <%s>'%(ponente.user.usuario.first_name,ponente.user.usuario.last_name,ponente.user.usuario.email),
            'activo':activo})
        context['ponentes_alls']=ponentes_env
        context['ponentes']=relaciones
        if self.request.GET.get('congreso'):
            context['con']=Congreso.objects.filter(path=self.request.GET.get('congreso')).first()
            context['blo']=Bloque.objects.filter(congreso=context['con'])
        if self.request.GET.get('bloque'):
            context['bloque']=Bloque.objects.filter(path=self.request.GET.get('bloque')).first()
            context['congreso']=context['bloque'].congreso
            context['blo']= None
        if self.request.GET.get('congreso_bloque'):
            context['congreso_bloque']=True
        return context 
      

    def form_valid(self, form):
        if not self.request.POST.getlist('taller_ponente-ponente'):
            messages.warning(self.request, 'Debe al menos entrar un ponente')
            return super().form_invalid(form) 
        taller_update=Taller.objects.get(pk=self.kwargs.get('pk'))
        self.object=taller_update
        taller=form['taller'].save(commit=False)
        ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)

        if ubic.exists():
            taller.lugar=ubic.first()
        else:
            ubicacion=Ubicacion(direccion=form['ubicacion'].instance.direccion,latitud=form['ubicacion'].instance.latitud,longitud=form['ubicacion'].instance.longitud)
            ubicacion.save()
            taller.lugar=ubicacion
        id_video=['']
        if taller.cod_video:
            id_video=taller.cod_video.split(sep='https://player.vimeo.com/video/')
            id_video=id_video[-1].split(sep='"')
        taller.id_video=id_video[0]
        imagen_seg=self.request.POST['taller-prueba']
        if 'talleres/' not in imagen_seg:
            image_64_encode=self.request.POST['taller-prueba']
            campo = image_64_encode.split(",")
            chars = '0123456789'
            nombre = get_random_string(5, chars)
            image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
            image_result = open('MedCongressApp/static/talleres/imagen_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
            image_result.write(image_64_decode)
            if taller.imagen:
                fileObj = Path('MedCongressApp/static/%s'%( taller.imagen))
                if fileObj.is_file():
                    remove('MedCongressApp/static/%s'%( taller.imagen))
            taller.imagen='talleres/imagen_%s.png'%(nombre)
        if self.request.POST['taller-constancia']:
            constancia=self.request.POST['taller-constancia']
            if 'congreso/' not in constancia:
            
                image_64_encode=self.request.POST['taller-constancia']
                campo = image_64_encode.split(",")
                chars = '0123456789'
                nombre = get_random_string(5, chars)
                image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8'))
                image_result = open('MedCongressApp/static/congreso/img_constancia/taller_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
                image_result.write(image_64_decode)
                if  taller_update.foto_constancia:
                    fileObj = Path('MedCongressApp/static/%s'%( taller_update.foto_constancia))
                    if fileObj.is_file():    
                        remove('MedCongressApp/static/%s'%( taller_update.foto_constancia))
                taller.foto_constancia='congreso/img_constancia/taller_%s.png'%(nombre)   
        taller_update=taller
        taller_update.save()
        relaciones=RelTallerPonente.objects.filter(taller=taller_update)
        relaciones.delete()
        for ponente in self.request.POST.getlist('taller_ponente-ponente'):
            ponente_=Ponente.objects.get(pk=ponente)
            po= RelTallerPonente(ponente=ponente_,taller=taller)
            po.save()
        return super(TallerUpdateView, self).form_valid(form)
    def get_success_url(self):
        url=reverse_lazy('MedCongressAdmin:talleres_list')
        self.success_url='%s?&search=%s'%(url,self.request.GET.get('search'))
        if self.request.GET.get('congreso'):
            self.success_url =  '%s?congreso=%s&search=%s'%(url,self.request.GET.get('congreso'),self.request.GET.get('search')) 
        if self.request.GET.get('bloque'): 
            self.success_url =  '%s?bloque=%s&search=%s'%(url,self.request.GET.get('bloque'),self.request.GET.get('search')) 
            if self.request.GET.get('congreso_bloque'):
                self.success_url =  '%s?bloque=%s&search=%s&congreso_bloque=true'%(url,self.request.GET.get('bloque'),self.request.GET.get('search')) 

        return self.success_url  

class TallerPonenteListView(validarOrganizador,TemplateView):
    template_name= 'MedCongressAdmin/taller/listar_ponentes.html' 
    def get(self, request, **kwargs):
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        if taller is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=taller.congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        return self.render_to_response(self.get_context_data())    
    def get_context_data(self, **kwargs):
        context = super(TallerPonenteListView, self).get_context_data(**kwargs)
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        context['taller']=taller
        context['ponentes']=RelTallerPonente.objects.filter(taller=taller)
        if self.request.GET.get('congreso'):
            context['congreso']=taller.congreso   
        if self.request.GET.get('bloque'):
            context['bloque']=taller.bloque
        if self.request.GET.get('congreso_bloque'):
            context['congreso']=taller.congreso
            context['bloque']=taller.bloque
            context['congreso_bloque']=True
        return context     

class TallerPonenteCreateView(validarOrganizador,CreateView):
    
    form_class = PonenteTallerForm
    # success_url = reverse_lazy('MedCongressAdmin:ponencias_list')
    template_name = 'MedCongressAdmin/taller/ponente_form.html'

    def get(self, request, **kwargs):
       
        taller=Taller.objects.filter(path=self.kwargs.get('path')).first()
        if taller is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=taller.congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        self.object=taller
        return self.render_to_response(self.get_context_data())   
    
    def form_valid(self, form):
        ponencia=form.save(commit=False)
  
        ponencia.save()
        return super(TallerPonenteCreateView, self).form_valid(form)

    def get_success_url(self):
        
        url =  reverse_lazy('MedCongressAdmin:Taller_ponentes',kwargs={'path': self.kwargs.get('path')} )+'?search=%s'%(self.request.GET.get('search'))
        if self.request.GET.get('congreso'):
            self.success_url =  '%s&congreso=True'%(url) 
        if self.request.GET.get('bloque'): 
            self.success_url =  '%s&bloque=True'%(url) 
        if self.request.GET.get('congreso_bloque'):
            self.success_url =  '%s&congreso_bloque=True'%(url) 
        return self.success_url 

   

    def get_context_data(self, **kwargs):
        ctx = super(TallerPonenteCreateView, self).get_context_data(**kwargs)
        pon=Taller.objects.filter(path=self.kwargs.get('path')).first()
        ctx['pon'] = pon
        ponentes=RelTallerPonente.objects.filter(taller=pon)
        id=[]
        for ponente in ponentes:
            id.append(ponente.ponente.pk)
        ctx['ponentes']=Ponente.objects.exclude(id__in=id)
        if self.request.GET.get('congreso'):
            ctx['congreso']=pon.congreso
        if self.request.GET.get('bloque'):
            ctx['bloque']=pon.bloque
        if self.request.GET.get('congreso_bloque'):
            ctx['congreso']=pon.congreso
            ctx['bloque']=pon.bloque
            ctx['congreso_bloque']=True 
        ctx['search']=self.request.GET.get('search')   
        return ctx


class TallerDeletedView(validarUser,DeleteView):
    model = Taller
    success_url = reverse_lazy('MedCongressAdmin:talleres_list')

class AsignarTalleresListView(validarUser,ListView,FormView):
    model = RelTallerUser
    context_object_name = 'talleres'
    template_name = 'MedCongressAdmin/asignar_taller.html'
    form_class=ExportarTallerExelForm

    def form_valid(self, form):
        self.object_list = self.get_queryset()
        #Obtenemos todas las personas de nuestra base de datos
        taller=self.request.POST['taller']
        query= RelTallerUser.objects.filter(taller=taller,is_pagado=True).values('user__usuario__first_name','user__usuario__last_name','user__usuario__email','user__genero__denominacion','categoria_pago__nombre','user__cel_profecional','user__categoria__nombre','user__ubicacion__direccion','user__especialidad__nombre','user__fecha_nacimiento','user__num_telefono','taller__titulo').annotate(Sum('cantidad'))
        
        if query:
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=40
            ws.column_dimensions['C'].width=40
            ws.column_dimensions['D'].width=47

            ws.column_dimensions['E'].width=20
            ws.column_dimensions['F'].width=20
            ws.column_dimensions['G'].width=27
            ws.column_dimensions['H'].width=56
            ws.column_dimensions['I'].width=20
            ws.column_dimensions['J'].width=25
            ws.column_dimensions['K'].width=25
            ws.column_dimensions['L'].width=22

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Usuarios que han comprado el  Taller :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            ws['A2'] ='" %s "'%(query[0]['taller__titulo']) 
            ws['A2'].font = Font(size=12,bold=True)
            ws['A2'].alignment = Alignment(mergeCell='center',horizontal='center') 
           
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'].style =titulo
            ws['B3'].style =titulo 
            ws['C3'].style =titulo 
            ws['D3'].style =titulo                   
            ws['E3'].style =titulo 
            ws['F3'].style =titulo 
            ws['G3'].style =titulo 
            ws['H3'].style =titulo
            ws['I3'].style =titulo 
            ws['J3'].style =titulo  
            ws['K3'].style =titulo 
            ws['L3'].style =titulo

            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Dirección'
            ws['E3'] = 'Teléfono'
            ws['F3'] = 'Género'
            ws['G3'] = 'Categoría'
            ws['H3'] = 'Especialidad'
            ws['I3'] = 'Cédula Profecional' 
            ws['J3'] = 'Fecha de Nacimiento' 
            ws['K3'] = 'Categoría de Pago'  
            ws['L3'] = 'Cantidad Comprados'           
            cont=4
            
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                
                ws.cell(row=cont,column=1).style=celdas
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).style=celdas
                ws.cell(row=cont,column=2).value ='%s %s'%(quer['user__usuario__first_name'],quer['user__usuario__last_name']) 
                ws.cell(row=cont,column=3).style=celdas
                ws.cell(row=cont,column=3).value = quer['user__usuario__email']
                ws.cell(row=cont,column=4).style=celdas
                ws.cell(row=cont,column=4).value = quer['user__ubicacion__direccion']
                ws.cell(row=cont,column=5).style=celdas
                ws.cell(row=cont,column=5).value = quer['user__num_telefono']
                ws.cell(row=cont,column=6).style=celdas
                ws.cell(row=cont,column=6).value = quer['user__genero__denominacion']
                ws.cell(row=cont,column=7).style=celdas
                ws.cell(row=cont,column=7).value = quer['user__categoria__nombre']
                ws.cell(row=cont,column=8).style=celdas
                ws.cell(row=cont,column=8).value = quer['user__especialidad__nombre']
                ws.cell(row=cont,column=9).style=celdas
                ws.cell(row=cont,column=9).value = quer['user__cel_profecional']
                ws.cell(row=cont,column=10).style=celdas
                ws.cell(row=cont,column=10).value = quer['user__fecha_nacimiento']
                ws.cell(row=cont,column=11).style=celdas
                ws.cell(row=cont,column=11).value = quer['categoria_pago__nombre']
                ws.cell(row=cont,column=12).style=celdas
                ws.cell(row=cont,column=12).value = quer['cantidad__sum']
                cont = cont + 1
            
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=RelTallerUser.xlsx"
            wb.save(response)
            return response
        else:
            taller=Taller.objects.get(pk=taller)
            messages.warning(self.request, 'Todavía ningún usuario ha comprado este Taller')
            return HttpResponseRedirect(reverse_lazy('MedCongressAdmin:asig_talleres_list')+'?exportar=%s'%(taller.path))


    def get_context_data(self, **kwargs):
        context = super(AsignarTalleresListView, self).get_context_data(**kwargs)
        context['search']=self.request.GET.get('search')
        if self.request.GET.get('exportar'):
            taller_evn=[]
            talleres= Taller.objects.all()
            activo=False
            for taller in talleres:
                if taller.path == self.request.GET.get('exportar'):
                    activo=True
                else:
                    activo=False
                taller_evn.append({ 'id':taller.pk,
                                    'titulo':taller.titulo,
                                        'activo':activo,

                })
            context['exportar']= taller_evn
            
        return context  

class AsignarTallerAddViews(validarUser,FormView):
    form_class = AsignarTallerForms
    success_url = reverse_lazy('MedCongressAdmin:asig_talleres_list')
    template_name = 'MedCongressAdmin/asig_taller_form.html'

    def form_valid(self, form):
        taller=form.save(commit=True)
        
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        ctx = super(AsignarTallerAddViews, self).get_context_data(**kwargs)
        ctx['search']=self.request.GET.get('search')
        return ctx 
    def get_success_url(self):
        url=reverse_lazy('MedCongressAdmin:asig_talleres_list')
        self.success_url='%s?search=%s'%(url,self.request.GET.get('search'))
        return self.success_url       

def GetPagosT(request):
    if request.is_ajax():
        query = request.POST['taller_id']
        categoria=RelTalleresCategoriaPago.objects.filter(taller=Taller.objects.get(pk=query))
        
        results = []
        for cat in categoria:
            results.append({'nombre':cat.categoria.nombre,'id':cat.categoria.pk,'moneda':cat.moneda})
            data = json.dumps(results)
    mimetype = "application/json"
    return HttpResponse(data, mimetype) 

class AsignarTallerDeletedViews(validarUser,DeleteView):
    model = RelTallerUser
    success_url = reverse_lazy('MedCongressAdmin:asig_talleres_list')

class TallerPonenteDeletedView(validarOrganizador,DeleteView):
    model = RelTallerPonente
    success_url = reverse_lazy('MedCongressAdmin:talleres_list')

def TallerBloqueDeleted(request):
    if request.is_ajax():
        query = request.POST['taller_id']
        taller=Taller.objects.get(id=query)
        taller.bloque=None
        taller.save()
        data = json.dumps([{'titulo':taller.titulo}])
        
        mimetype = "application/json"
    return HttpResponse(data, mimetype)   
    
class ReporteRelTallerUserExcel(TemplateView):
    
    #Usamos el método get para generar el archivo excel 
    def post(self, request):
        #Obtenemos todas las personas de nuestra base de datos
        taller=self.request.POST['taller']
        query= RelTallerUser.objects.filter(taller=taller,is_pagado=True).values('user__usuario__first_name','user__usuario__last_name','user__usuario__email','taller__titulo','categoria_pago__nombre').annotate(Sum('cantidad'))
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        if query:
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['B1'] = 'Usuarios que han comprado el Taller%s'%( query[0]['taller__titulo'])
            ws['B1'].font = Font(size=12,bold=True)
            ws['B1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('B1:E1')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Taller'
            ws['E3'] = 'Categoria de Pago'
            ws['F3'] = 'Cantidad'        
            cont=4
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).value = '%s %s'%(quer['user__usuario__first_name'],quer['user__usuario__last_name'])
                ws.cell(row=cont,column=3).value = quer['user__usuario__email']
                ws.cell(row=cont,column=4).value = quer['taller__titulo']
                ws.cell(row=cont,column=5).value = quer['categoria_pago__nombre']
                ws.cell(row=cont,column=6).value = quer['cantidad__sum']
                cont = cont + 1
        else:
            ws['B1'] = 'Este Taller nadie lo ha comprado aún'
            ws['B1'].font = Font(size=12,bold=True)
            ws['B1'].alignment = Alignment(mergeCell='center',horizontal='center') 
        response = HttpResponse(content_type="application/ms-excel") 
        response["Content-Disposition"] = "attachment; filename=RelTallerUser.xlsx"
        wb.save(response)
        return response

class AsignarConstanciasTaller(validarUser,TemplateView):
    template_name = 'MedCongressAdmin/asig_constancia_taller.html'
    def get_context_data(self, **kwargs):
        context = super(AsignarConstanciasTaller, self).get_context_data(**kwargs)
        congreso=Taller.objects.all()
        context['congresos']=congreso
        return context 
    def post(self, request, **kwargs):
        titulo= self.request.POST['my_congress']
        taller=Taller.objects.get(pk=self.request.POST['my_congress'])
        if taller.foto_constancia:
            if taller:
                prueba=Constanciataller.apply_async(args=[titulo])
                messages.warning(self.request,'Se creo la constancia a todos los que participaron el Taller %s'%(taller.titulo))
                return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_taller'))
            else:
                messages.warning(self.request,'Ese Taller no existe')
                return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_taller'))
        else:
                messages.warning(self.request,'Error.....Ese Taller tiene asignada ninguna foto para la constancia')
                return HttpResponseRedirect(reverse('MedCongressAdmin:asig_constancia_taller'))

class TallerCategPagosUpdateView(validarOrganizador,UpdateView):

    form_class = TallerCategPagoForm
    success_url = reverse_lazy('MedCongressAdmin:taller_list')
    template_name = 'MedCongressAdmin/taller/cat_pago_form.html'

    def get_queryset(self, **kwargs):
        return RelTalleresCategoriaPago.objects.filter(pk=self.kwargs.get('pk'))

    def get(self, request, **kwargs):
        taller_cat=RelTalleresCategoriaPago.objects.filter(pk=self.kwargs.get('pk')).first()
        if taller_cat is None:
            return   HttpResponseRedirect(reverse('Error404'))
        if not Organizador.objects.filter(user=self.request.user.perfilusuario,congreso=taller_cat.taller.congreso).exists() and not self.request.user.is_staff: 
            return   HttpResponseRedirect(reverse('Error403'))
        self.object=taller_cat
        return self.render_to_response(self.get_context_data()) 

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['update']=True
        context['taller'] = self.object.taller
        if self.request.GET.get('congreso'):
            context['congreso']=self.object.taller.congreso   
        if self.request.GET.get('bloque'):
            context['bloque']=self.object.taller.bloque
        if self.request.GET.get('congreso_bloque'):
            context['congreso']=self.object.taller.congreso
            context['bloque']=self.object.taller.bloque
            context['congreso_bloque']=True

        return context

    def get_success_url(self):
        url =  reverse_lazy('MedCongressAdmin:Taller_pagos',kwargs={'path': self.object.taller.path} )+'?&search=%s'%(self.request.GET.get('search'))
        if self.request.GET.get('congreso'):
            url =  '%s&congreso=True'%(url) 
        if self.request.GET.get('bloque'): 
            url =  '%s&bloque=True'%(url) 
        if self.request.GET.get('congreso_bloque'):
            url =  '%s&congreso_bloque=True'%(url) 
        self.success_url=url
        return self.success_url

class TallerCategPagosDeletedView(validarOrganizador,DeleteView):
    model = RelTalleresCategoriaPago
    success_url = reverse_lazy('MedCongressAdmin:cat_usuarios_list')

class vTableAsJSONTaller(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['titulo','congreso__titulo','','published']
           
        #listado que muestra en dependencia de donde estes parado
       
        if request.GET.get('tipo')=='nada':
            object_list = Taller.objects.all()
        if request.GET.get('tipo')=='congreso':
            object_list = Taller.objects.filter(congreso__path=request.GET.get('path'))
        if request.GET.get('tipo')=='bloque':
            object_list = Taller.objects.filter(bloque__path=request.GET.get('path'))
        
        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        if sort_col!=4 or sort_col!=2 :# columna en la tabla para las operaciones
            sort_colr = col_name_map[sort_col]
            object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(titulo__icontains=search_text) | Q(congreso__titulo__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
        for objet in filtered_object_list[start:(start+delta)]:
            public='No'
            if objet.published:
                public='Si'
            operaciones=''
            operaciones=''' <a href="'''+ reverse('MedCongressAdmin:taller_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>'''
            if request.GET.get('tipo')=='nada':
                operaciones=''' <a href="'''+ reverse('MedCongressAdmin:taller_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>'''
            if request.GET.get('tipo')=='congreso':
                
                operaciones=''' <a href="'''+ reverse('MedCongressAdmin:taller_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''&congreso='''+request.GET.get('path')+'''"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>'''
            if request.GET.get('tipo')=='bloque':
                if request.GET.get('congreso_bloque'):
                    operaciones=''' <a href="'''+ reverse('MedCongressAdmin:taller_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''&bloque='''+request.GET.get('path')+'''&congreso_bloque=true"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>'''
                else:
                    operaciones=''' <a href="'''+ reverse('MedCongressAdmin:taller_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''&bloque='''+request.GET.get('path')+'''"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>'''
                
            
            ponentes=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_ponentes',kwargs={'path':objet.path})+'''"
                                                        title="Ponentes">
                                                        <i class="icon icon-ponente " style= "color: blue;" ></i>
                                                    </a>'''
            if request.GET.get('tipo')=='congreso':
                ponentes=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_ponentes',kwargs={'path':objet.path})+'''?congreso=True"
                                                        title="Ponentes">
                                                        <i class="icon icon-ponente " style= "color: blue;" ></i>
                                                    </a>'''
                
                
            if request.GET.get('tipo')=='bloque':
                if request.GET.get('congreso_bloque'):
                    ponentes=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_ponentes',kwargs={'path':objet.path})+'''?congreso_bloque=true"
                                                        title="Ponentes">
                                                        <i class="icon icon-ponente " style= "color: blue;" ></i>
                                                    </a>'''
                    
                else:
                    ponentes=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_ponentes',kwargs={'path':objet.path})+'''?bloque=True"
                                                        title="Ponentes">
                                                        <i class="icon icon-ponente " style= "color: blue;" ></i>
                                                    </a>'''
                    
            cat_pago=''
            cat_pago=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_pagos',kwargs={'path':objet.path})+'''"
                                                        title="Categorias de pago">
                                                        <i class="icon icon-pago " style= "color: blue;" ></i>
                                                    </a>'''
            if request.GET.get('tipo')=='nada':
                cat_pago=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_pagos',kwargs={'path':objet.path})+'''"
                                                        title="Categorias de pago">
                                                        <i class="icon icon-pago " style= "color: blue;" ></i>
                                                    </a>'''
            if request.GET.get('tipo')=='congreso':
                cat_pago=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_pagos',kwargs={'path':objet.path})+'''?congreso=True"
                                                        title="Categorias de pago">
                                                        <i class="icon icon-pago " style= "color: blue;" ></i>
                                                    </a>'''
                
                
            if request.GET.get('tipo')=='bloque':
                if request.GET.get('congreso_bloque'):
                    cat_pago=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_pagos',kwargs={'path':objet.path})+'''?congreso_bloque=true"
                                                        title="Categorias de pago">
                                                        <i class="icon icon-pago " style= "color: blue;" ></i>
                                                    </a>'''
                    
                else:
                    cat_pago=''' <a  href="'''+ reverse('MedCongressAdmin:Taller_pagos',kwargs={'path':objet.path})+'''?bloque=True"
                                                        title="Categorias de pago">
                                                        <i class="icon icon-pago " style= "color: blue;" ></i>
                                                    </a>'''
                      
           #Guardar datos en un dic 
            
            enviar.append({ 'nombre':objet.titulo,
                            'congreso': objet.congreso.titulo,
                            'ponentes':ponentes,
                            'public' : public,
                            'cat_pago':cat_pago,
                            'operaciones' : operaciones,
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)

class vTableAsJSONAsigTaller(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['user__usuario__first_name','user__usuario__email','taller__titulo','cantidad','categoria_pago__nombre','is_pagado']
           
        #listado que muestra en dependencia de donde estes parado
        object_list = RelTallerUser.objects.all()
        
        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        
        sort_colr = col_name_map[sort_col]
        object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(user__usuario__last_name__icontains=search_text) | Q(user__usuario__email__icontains=search_text)|Q(user__usuario__first_name__icontains=search_text)|Q(taller__titulo__icontains=search_text)|Q(cantidad__icontains=search_text)|Q(categoria_pago__nombre__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
       
            # if objet.ponente:
            #     user= '%s %s'%(objet.ponente.first().user.usuario.first_name,objet.ponente.first().user.usuario.last_name)
           
           #Guardar datos en un dic 
        for objet in filtered_object_list[start:(start+delta)]:
            pagado='No'
            if objet.is_pagado:
                pagado='Si'
                
            enviar.append({ 'usuario':'%s %s'%(objet.user.usuario.first_name,objet.user.usuario.last_name),
                            'email': objet.user.usuario.email,
                            'congreso' : objet.taller.titulo,
                            'cantidad' : objet.cantidad,
                            'cat_pago':objet.categoria_pago.nombre,
                            'pagado':pagado,
                            'operaciones' : 
                                                    '''<a id="del_'''+ str(objet.pk)+'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk)+''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>''',
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)  