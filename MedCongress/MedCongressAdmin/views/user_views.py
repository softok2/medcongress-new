import base64
import json
from datetime import datetime, timedelta,timezone
from os import remove
from pathlib import Path

from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import AccessMixin, UserPassesTestMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import (HttpResponse, HttpResponseBadRequest,
                         HttpResponseRedirect)
from django.shortcuts import redirect
from django.urls import reverse, reverse_lazy
from django.utils.crypto import get_random_string
from django.views.generic import CreateView, ListView, TemplateView
from django.views.generic.edit import DeleteView, FormView, UpdateView
from MedCongressAdmin.apps import validarUser
from MedCongressAdmin.forms.congres_forms import (ExportarLogsUsuarioExelForm,
                                                  UsuarioForms)
from MedCongressApp.models import (CategoriaUsuario, Especialidades,
                                   PerfilUsuario, Ubicacion, User,UserActivityLog)
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Protection, Side)

# from django.views import generic
  


class UsuariosListView(validarUser,ListView):
    model = PerfilUsuario
    context_object_name = 'users'
    template_name = 'MedCongressAdmin/usuarios.html'
  
    def get_context_data(self, **kwargs):
        context=super(UsuariosListView,self).get_context_data(**kwargs)
        context['search']=self.request.GET.get('search')
        context['categorias']=CategoriaUsuario.objects.all().distinct('nombre').order_by('nombre')
        context['especialidades']=Especialidades.objects.all().order_by('nombre')
        return context
    def post(self, request, **kwargs):
        if self.request.POST.getlist('categoria') and self.request.POST.getlist('especialidad'):
            categorias=self.request.POST.getlist('categoria')
            especialidades=self.request.POST.getlist('especialidad')
            query= PerfilUsuario.objects.filter(categoria__in=categorias,especialidad__in=especialidades).order_by('categoria__nombre','especialidad__nombre')
        elif self.request.POST.getlist('especialidad'):
            especialidades=self.request.POST.getlist('especialidad')
            query= PerfilUsuario.objects.filter(especialidad__in=especialidades).order_by('categoria__nombre','especialidad__nombre') 
        elif self.request.POST.getlist('categoria'):
            categorias=self.request.POST.getlist('categoria')
            query= PerfilUsuario.objects.filter(categoria__in=categorias).order_by('categoria__nombre','especialidad__nombre')
        else:
            query= PerfilUsuario.objects.all().order_by('categoria__nombre','especialidad__nombre')
        if query:
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=40
            ws.column_dimensions['C'].width=40
            ws.column_dimensions['D'].width=47

            ws.column_dimensions['E'].width=20
            ws.column_dimensions['F'].width=20
            ws.column_dimensions['G'].width=27
            ws.column_dimensions['H'].width=56
            ws.column_dimensions['I'].width=20
            ws.column_dimensions['J'].width=25
            ws.column_dimensions['K'].width=25
            ws.column_dimensions['L'].width=22

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Usuarios registrados :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 

            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'].style =titulo
            ws['B3'].style =titulo 
            ws['C3'].style =titulo 
            ws['D3'].style =titulo                   
            ws['E3'].style =titulo 
            ws['F3'].style =titulo 
            ws['G3'].style =titulo 
            ws['H3'].style =titulo
            ws['I3'].style =titulo 
            ws['J3'].style =titulo  
            ws['K3'].style =titulo 
            ws['L3'].style =titulo

            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Dirección'
            ws['E3'] = 'Teléfono'
            ws['F3'] = 'Género'
            ws['G3'] = 'Categoría'
            ws['H3'] = 'Especialidad'
            ws['I3'] = 'Cédula Profecional' 
            ws['J3'] = 'Fecha de Nacimiento' 
          
            cont=4
            
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                
                ws.cell(row=cont,column=1).style=celdas
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).style=celdas
                ws.cell(row=cont,column=2).value ='%s %s'%(quer.usuario.first_name,quer.usuario.last_name) 
                ws.cell(row=cont,column=3).style=celdas
                ws.cell(row=cont,column=3).value = quer.usuario.email
                ws.cell(row=cont,column=4).style=celdas
                if quer.ubicacion:
                    ws.cell(row=cont,column=4).value = quer.ubicacion.direccion
                else:
                    ws.cell(row=cont,column=4).value = ' '
                ws.cell(row=cont,column=5).style=celdas
                ws.cell(row=cont,column=5).value = quer.num_telefono
                ws.cell(row=cont,column=6).style=celdas
                if quer.genero:
                    ws.cell(row=cont,column=6).value = quer.genero.denominacion
                else:
                    ws.cell(row=cont,column=6).value = ' '
                ws.cell(row=cont,column=7).style=celdas
                ws.cell(row=cont,column=7).value = quer.categoria.nombre
                ws.cell(row=cont,column=8).style=celdas
                if quer.especialidad:
                    ws.cell(row=cont,column=8).value = quer.especialidad.nombre
                else:
                    ws.cell(row=cont,column=8).value = ''
                ws.cell(row=cont,column=9).style=celdas
                ws.cell(row=cont,column=9).value = quer.cel_profecional
                ws.cell(row=cont,column=10).style=celdas
                ws.cell(row=cont,column=10).value = quer.fecha_nacimiento
               
                cont = cont + 1
            
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=Usuarios.xlsx"
            wb.save(response)
            return response
        else:
            messages.warning(self.request, 'No hay Usuarios con ese criterio de búsqueda')
            return HttpResponseRedirect(reverse_lazy('MedCongressAdmin:usuarios_list'))

class UsuarioCreateView(validarUser,FormView):
    model=User
    form_class = UsuarioForms
    success_url = reverse_lazy('MedCongressAdmin:usuarios_list')
    template_name = 'MedCongressAdmin/usuario_form.html'
    def get_success_url(self):
        url= reverse_lazy('MedCongressAdmin:usuarios_list')
        self.success_url='%s?&search=%s'%(url,self.request.GET.get('search'))
        return self.success_url  
    def form_valid(self, form):
        user = form['user'].save(commit=False)
        perfiluser = form['perfiluser'].save(commit=False)
        ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)
        
        if ubic.exists():
            perfiluser.ubicacion=ubic.first()
        else:
            ubicacion=form['ubicacion'].save(commit=True)
            perfiluser.ubicacion=ubicacion
       
        us=User.objects.create_user(user.email,user.email,user.password)  
        us.first_name=user.first_name
        us.last_name=user.last_name
        us.is_active = True
        us.save()
        perfiluser.usuario = us
        perfiluser.path=us.username
        if self.request.POST['prueba']:
            image_64_encode=self.request.POST['prueba']
            campo = image_64_encode.split(",")
            image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
            nombre=user.email.replace('.','_')
            image_result = open('MedCongressApp/static/usuarios/foto_%s.png'%(nombre), 'wb') # create a writable image and write the decoding result
            image_result.write(image_64_decode)
            perfiluser.foto='usuarios/foto_%s.png'%(nombre)
        else:
            perfiluser.foto='usuarios/defaulthombre.png'
        perfiluser.save() 
        return super(UsuarioCreateView, self).form_valid(form)

class UsuarioUpdateView(validarUser,FormView):
    form_class = UsuarioForms
    success_url = reverse_lazy('MedCongressAdmin:usuarios_list')
    template_name = 'MedCongressAdmin/usuario_form.html'

    def get_queryset(self, **kwargs):
        return PerfilUsuario.objects.filter(pk=self.kwargs.get('pk'))
    
    def get_form_kwargs(self):
        kwargs = super(UsuarioUpdateView, self).get_form_kwargs()
        self.object=PerfilUsuario.objects.get(pk=self.kwargs.get('pk'))
        kwargs.update(instance={
            'perfiluser': self.object,
            'user': self.object.usuario,
            'ubicacion': self.object.ubicacion,
        })
        return kwargs

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        if self.object.meta_og_imagen:
            context['imagen_meta']='/static/%s'%(self.object.meta_og_imagen)
        if self.object.foto:    
            context['imagen_seg_url']=self.object.foto
        context['update']=True
        context['usuario']=self.object
        return context

    def form_valid(self, form):
        perfiluser_update=PerfilUsuario.objects.get(pk=self.kwargs.get('pk'))
        user_update=perfiluser_update.usuario
        user = form['user'].save(commit=False)
        perfiluser = form['perfiluser'].save(commit=False)
        user_update=user
        perfiluser_update=perfiluser
        user_update.save()
        ubic=Ubicacion.objects.filter(direccion=form['ubicacion'].instance.direccion)
        
        if ubic.exists():
            perfiluser_update.ubicacion=ubic.first()
        else:
            ubicacion=Ubicacion(direccion=form['ubicacion'].instance.direccion,latitud=form['ubicacion'].instance.latitud,longitud=form['ubicacion'].instance.longitud)
            ubicacion.save()
            perfiluser_update.ubicacion=ubicacion

        perfiluser_update.usuario = user_update
        if self.request.POST['prueba']:
            image_64_encode=self.request.POST['prueba']
            campo = image_64_encode.split(",")
            image_64_decode = base64.decodestring(bytes(campo[1], encoding='utf8')) 
            nombre=user.email.replace('.','_')
            chars = '0123456789'
            nom= get_random_string(5, chars)
            image_result = open('MedCongressApp/static/usuarios/foto_%s_%s.png'%(nombre,nom), 'wb') # create a writable image and write the decoding result
            image_result.write(image_64_decode)
            if perfiluser.foto and perfiluser.foto!= 'usuarios/defaulthombre.png' :
                fileObj = Path('MedCongressApp/static/%s'%( perfiluser.foto))
                if fileObj.is_file():    
                    remove('MedCongressApp/static/%s'%( perfiluser.foto))
            perfiluser.foto='usuarios/foto_%s_%s.png'%(nombre,nom)
        else:
            if not perfiluser.foto :
                perfiluser.foto='usuarios/defaulthombre.png'
        perfiluser_update.save() 
        return super(UsuarioUpdateView, self).form_valid(form)
    def get_success_url(self):
        url= reverse_lazy('MedCongressAdmin:usuarios_list')
        self.success_url='%s?search=%s'%(url,self.request.GET.get('search'))
        return self.success_url  
class UsuarioDeletedView(validarUser,DeleteView):
    model = User
    success_url = reverse_lazy('MedCongressAdmin:usuarios_list')

class UsuarioAsigCongresoUserView(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'

    def get_context_data(self, **kwargs):
        ctx = super(UsuarioAsigCongresoUserView, self).get_context_data(**kwargs)
        usuario=PerfilUsuario.objects.get(pk=self.kwargs.get('pk'))
        ctx['usuario'] = usuario
        return ctx

class vTableAsJSON(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['usuario__first_name','usuario__email','categoria__nombre','especialidad__nombre']
           
        #listado que muestra en dependencia de donde estes parado
        object_list = PerfilUsuario.objects.all()
        
        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        if sort_col!=4 or sort_col!=2 :# columna en la tabla para las operaciones
            sort_colr = col_name_map[sort_col]
            object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(usuario__last_name__icontains=search_text) | Q(usuario__email__icontains=search_text)|Q(usuario__first_name__icontains=search_text)|Q(especialidad__nombre__icontains=search_text)|Q(categoria__nombre__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
       
            # if objet.ponente:
            #     user= '%s %s'%(objet.ponente.first().user.usuario.first_name,objet.ponente.first().user.usuario.last_name)
           
           #Guardar datos en un dic 
        for objet in filtered_object_list[start:(start+delta)]:
            especialidad=''
            if objet.especialidad:
                especialidad=objet.especialidad.nombre
            enviar.append({ 'nombre':'%s %s'%(objet.usuario.first_name,objet.usuario.last_name),
                            'email': objet.usuario.email,
                            'categoria' : objet.categoria.nombre,
                            'especialidad' : especialidad,
                            'operaciones' : ''' <a href="'''+ reverse('MedCongressAdmin:usuario_edit',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''"
                                                    title="Editar"><i class="icon icon-editar"></i></a>
                                                    <a href="'''+ reverse('MedCongressAdmin:asig_congreso',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''&ponente=True"
                                                    title="Asignar Congreso"><i class="icon icon-asignar_congreso"></i></a>
                                                    <a id="del_'''+ str(objet.pk) +'''"
                                                        href="javascript:deleteItem('''+ str(objet.usuario.pk) +''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>''',
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)    

class LogsUsuarios(validarUser,FormView):
    form_class=ExportarLogsUsuarioExelForm
    template_name = 'MedCongressAdmin/usuario/log_form.html'
    def form_valid(self, form):
        id_usuario=self.request.POST['usuario']
        usuario=PerfilUsuario.objects.get(pk=id_usuario)
        if self.request.POST['fecha_fin']:
            fecha_fin=self.request.POST['fecha_fin']
            d_fecha_final = datetime.strptime(fecha_fin,'%Y-%m-%d') + timedelta(days=1)
        else:
             d_fecha_final = datetime.now(timezone.utc) + timedelta(days=1)   
        fecha_inicio=self.request.POST['fecha_inicio']
        if fecha_inicio:
            d_fecha_inicio = datetime.strptime(fecha_inicio,'%Y-%m-%d')
            query=UserActivityLog.objects.filter(user=id_usuario,fecha__range=[d_fecha_inicio,d_fecha_final]).order_by('congreso','fecha')
        else:
            query=UserActivityLog.objects.filter(user=id_usuario,fecha__lt=d_fecha_final).order_by('congreso','fecha')
        
        if query:
        #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=25
            ws.column_dimensions['C'].width=20
            ws.column_dimensions['D'].width=150

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            celdas_fecha = NamedStyle(name="celdas_fecha")
            celdas_fecha.font=Font(size=12)
            
            celdas_fecha.alignment=Alignment(horizontal='center',mergeCell=True)
            celdas_fecha.border = Border(left=Side(border_style='thin',
                            color='FF000000'),
                    right=Side(border_style='thin',
                            color='FF000000'),
                    top=Side(border_style='thin',
                            color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                color='FF000000'),
                    vertical=Side(border_style='thin',
                                color='FF000000'),
                    horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            label = NamedStyle(name="label")
            label.font=Font(size=12,bold=True)
            label.alignment=Alignment(horizontal='right',mergeCell=True)
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Logs del Usuario :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            ws['A2'] ='%s %s <<%s>>'%(usuario.usuario.first_name,usuario.usuario.last_name,usuario.usuario.email)
            ws['A2'].font = Font(size=12,bold=True)
            ws['A2'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            cont=3
            num=1
            id_congreso=0
            if query:
                for quer in query:
                    if id_congreso!=quer.congreso.pk:
                        cont = cont + 1
                        num=1
                        ws.merge_cells('A%s:B%s'%(cont,cont))
                        ws.merge_cells('C%s:D%s'%(cont,cont))
                        ws.cell(row=cont,column=1).style=label
                        ws['A%s'%(cont)] = 'Congreso:'
                        ws['C%s'%(cont)] = '%s'%(quer.congreso.titulo)
                        cont = cont + 1
                        ws.cell(row=cont,column=1).style=titulo
                        ws.cell(row=cont,column=2).style=titulo
                        ws.cell(row=cont,column=3).style=titulo
                        ws.cell(row=cont,column=4).style=titulo
                        
                        ws.cell(row=cont,column=1).value='No.'
                        ws.cell(row=cont,column=2).value='Fecha' 
                        ws.cell(row=cont,column=3).value='Tiempo (H:M:S)'
                        ws.cell(row=cont,column=4).value='Acción'
                        id_congreso= quer.congreso.pk   
                    
                        cont = cont + 1
                #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
                
                    
                    ws.cell(row=cont,column=1).style=celdas_fecha
                    ws.cell(row=cont,column=1).value = num
                    ws.cell(row=cont,column=2).style=celdas_fecha
                    ws.cell(row=cont,column=2).value = quer.fecha
                    ws.cell(row=cont,column=3).style=celdas_fecha
                    tiempo=quer.tiempo.split('.')
                    ws.cell(row=cont,column=3).value =tiempo[0]
                    ws.cell(row=cont,column=4).style=celdas
                    ws.cell(row=cont,column=4).value = '  %s'%(quer.mensaje) 
                    
                    cont = cont + 1
                    num = num + 1
            else:
                ws.merge_cells('A%s:F%s'%(cont,cont))
                ws.cell(row=cont,column=1).value = 'Este usuario no tiene Logs'
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=LogsUsuario.xlsx"
            wb.save(response)
            return response
        
        else:
            messages.warning(self.request, 'No existen Logs con estos criterios de busquedas')
            return HttpResponseRedirect(reverse('MedCongressAdmin:LogsUsuario'))
