import base64
import json
from os import remove
from pathlib import Path

import pandas as pd
from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin
from django.core.exceptions import RequestDataTooBig, ValidationError
from django.db.models import Q,Sum
from django.http import (HttpResponse, HttpResponseBadRequest,
                         HttpResponseRedirect, JsonResponse)
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.crypto import get_random_string
from django.views.defaults import page_not_found
from django.views.generic import CreateView, ListView, TemplateView
from django.views.generic.edit import DeleteView, FormView, UpdateView
from MedCongressAdmin.apps import validarOrganizador, validarUser
from MedCongressAdmin.forms.congres_forms import (BloqueForms,
                                                  ExportarExelForm,
                                                  ModeradorBloqueForm,
                                                  SelectPonencia)
from MedCongressAdmin.task import AsignarBeca
from MedCongressApp.models import (Bloque, Congreso, Moderador, Organizador,
                                   Ponencia, RelBloqueModerador,
                                   RelCongresoUser, Sala, Taller)
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Protection, Side)
from pandas import isnull


class vTableAsJSONBecaCongreso(TemplateView):
    template_name = 'MedCongressAdmin/asig_congress_form.html'
    
    def get(self, request, *args, **kwargs):
        #arreglo con las columnas de la BD a filtrar
        col_name_map = ['user__usuario__first_name','user__usuario__email','congreso__titulo','cantidad','categoria_pago__nombre','is_pagado','is_constancia']
           
        #listado que muestra en dependencia de donde estes parado

        if request.user.is_staff:
            object_list = RelCongresoUser.objects.filter(is_beca=True)
        else:
            congreso_orgs=Organizador.objects.filter(user=request.user.perfilusuario)
            object_list=RelCongresoUser.objects.filter(pk=0)
            for organizador in congreso_orgs:
                object_list |=RelCongresoUser.objects.filter(is_beca=True,congreso=organizador.congreso)
        
        
        #parametros 
        search_text = request.GET.get('sSearch', '').lower()# texto a buscar
        start = int(request.GET.get('iDisplayStart', 0))#por donde empezar a mostrar
        delta = int(request.GET.get('iDisplayLength', 10))#cantidad a mostrar
        sort_dir = request.GET.get('sSortDir_0', 'asc')# direccion a ordenar
        sort_col = int(request.GET.get('iSortCol_0', 0)) # numero de la columna a ordenar
        sort_col_name = request.GET.get('mDataProp_%s' % sort_col, '1')
        sort_dir_prefix = (sort_dir == 'desc' and '-' or '') #sufijo para poner en la consulta para ordenar

        #para ordenar el listado
        
        sort_colr = col_name_map[sort_col]
        object_list = object_list.order_by('%s%s' % (sort_dir_prefix,sort_colr))

        #para filtrar el listado
        filtered_object_list = object_list
        if len(search_text) > 0:
            filtered_object_list = object_list.filter(Q(user__usuario__last_name__icontains=search_text) | Q(user__usuario__email__icontains=search_text)|Q(user__usuario__first_name__icontains=search_text)|Q(congreso__titulo__icontains=search_text)|Q(cantidad__icontains=search_text)|Q(categoria_pago__nombre__icontains=search_text))

        #Guardar datos en un 
        enviar =[]
       
            # if objet.ponente:
            #     user= '%s %s'%(objet.ponente.first().user.usuario.first_name,objet.ponente.first().user.usuario.last_name)
           
           #Guardar datos en un dic 
        for objet in filtered_object_list[start:(start+delta)]:
           
           
            constancia='Si'
            
            if objet.foto_constancia or RelCongresoUser.objects.filter(congreso=objet.congreso,user=objet.user,is_constancia=True).exists():
                        constancia='''Si'''                                         
            else:
                constancia= '''<a href="'''+ reverse('MedCongressAdmin:constancia_usuario_add',kwargs={'pk':objet.pk})+'''?search='''+request.GET.get('search')+'''"
                                                    title="Asignar Constancia">
                                                    <i class="icon icon-constancia"></i>
                                                </a>'''   
           
            enviar.append({ 'usuario':'%s %s'%(objet.user.usuario.first_name,objet.user.usuario.last_name),
                            'email':'<p class="text"  >'+ objet.user.usuario.email+'</p>',
                            'congreso' : objet.congreso.titulo,
                            'cantidad' : objet.cantidad,
                            'cat_pago':'Beca',
                            'constancia':constancia,
                            'operaciones' : 
                                                    '''<a id="del_'''+ str(objet.pk)+'''"
                                                        href="javascript:deleteItem('''+ str(objet.pk)+''')"
                                                        title="Eliminar">
                                                        <i class="icon icon-eliminar"></i>
                                                    </a>''',
                            
            })
        #parametros para la respuesta
        jsoner = {
            
            "iTotalRecords": filtered_object_list.count(),
            "iTotalDisplayRecords": filtered_object_list.count(),
            "sEcho": request.GET.get('sEcho', 1),
            "data": enviar
        }
        data = json.dumps(jsoner)
        mimetype = "application/json"
        #Enviar
        return HttpResponse(data, mimetype)    

class BecasCongressListView(validarOrganizador,ListView):
    
    template_name = 'MedCongressAdmin/beca/listar.html'
   
    
    def post(self, request, **kwargs):
        try:
            prueba =True
            archivo=self.request.FILES['exel']
            if archivo:
                filename = archivo.name
            else:
                 raise ValidationError('Debe subir un Exel')   
            if(not filename.endswith(".xls") and not filename.endswith(".xlsx")):
                raise ValidationError('Debe subir un Exel')
               
            elif filename.endswith(".xls"):
                df = pd.read_excel(archivo)   
            else: 
                df = pd.read_excel(archivo, engine='openpyxl')
            df=df.applymap(lambda x: {} if isnull(x) else x)
            rows=df.to_dict('records')
            if not rows[0]['Correo'] or not rows[0]['Congreso']:
                raise ValidationError('Debe subir un Exel')
            res=AsignarBeca.delay(rows,self.request.user.pk)
            resultado=res.get()
            print(resultado)
            if resultado =='congreso':
                messages.warning(self.request, 'En este exel hay nombres de congreso que no existen en el sistema')
            if resultado =='usuario':
                messages.warning(self.request, 'En este exel hay correos que no son validos y no se guardaron en el sistema')
            if resultado =='no_permiso':
                
                messages.warning(self.request, 'En este exel hay congresos que no estas autorizado a asignar constancias')
            return HttpResponseRedirect(reverse('MedCongressAdmin:asig_becas_list'))
        except ValidationError as e:
            messages.warning(self.request, 'Debe entrar un archivo <b> EXEL (*.xls o *.xlsx)</b>')
            return HttpResponseRedirect(reverse('MedCongressAdmin:asig_becas_list'))
        except OSError :
            messages.warning(self.request, 'No está entrando los datos bien en el Exel')
            return HttpResponseRedirect(reverse('MedCongressAdmin:asig_becas_list'))
        except KeyError :
            messages.warning(self.request, 'No está entrando los datos bien en el Exel')
            return HttpResponseRedirect(reverse('MedCongressAdmin:asig_becas_list'))
        except ValueError:
            messages.warning(self.request, 'El tamaño de letra del exel debe ser menor de 14')
            return HttpResponseRedirect(reverse('MedCongressAdmin:asig_becas_list'))
    
    def get_queryset(self):
        
        queryset=RelCongresoUser.objects.filter(is_beca=True)
       
        return queryset
    def get_context_data(self, **kwargs):
        context = super(BecasCongressListView, self).get_context_data(**kwargs)
        if self.request.user.is_staff :
           
            congresos= Congreso.objects.all()
        else:
            congresos=[]
            org_congresos=Organizador.objects.filter(user=self.request.user.perfilusuario)
            for org_congreso in org_congresos:
                congresos.append(org_congreso.congreso)

        if self.request.GET.get('exportar'):
            congreso_evn=[]
           
            activo=False
            for congreso in congresos:
                if congreso.path == self.request.GET.get('exportar'):
                    activo=True
                else:
                    activo=False
               
                congreso_evn.append({ 'id':congreso.pk,
                                    'titulo':congreso.titulo,
                                        'activo':activo,

                })
            context['exportar']= congreso_evn
        
        context['congresos']= congresos
            
        return context  
 
class ExportarBecas(validarOrganizador,FormView):
    model = RelCongresoUser
    context_object_name = 'congress'
    template_name = 'MedCongressAdmin/beca/listar.html'
    form_class=ExportarExelForm
    
    def form_valid(self, form):
        self.object_list = self.get_queryset()
        id_congreso=self.request.POST['congreso']
        query= RelCongresoUser.objects.filter(congreso=id_congreso,is_pagado=True,is_beca=True).values('user__usuario__first_name','user__usuario__last_name','user__usuario__email','user__genero__denominacion','categoria_pago__nombre','user__cel_profecional','user__categoria__nombre','user__ubicacion__direccion','user__especialidad__nombre','user__fecha_nacimiento','user__num_telefono','congreso__titulo','user__puesto').annotate(Sum('cantidad'))
        
        if query:
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            ws.column_dimensions['A'].width=5
            ws.column_dimensions['B'].width=40
            ws.column_dimensions['C'].width=40
            ws.column_dimensions['D'].width=47

            ws.column_dimensions['E'].width=20
            ws.column_dimensions['F'].width=20
            ws.column_dimensions['G'].width=27
            ws.column_dimensions['H'].width=56
            ws.column_dimensions['I'].width=20
            ws.column_dimensions['J'].width=25
            ws.column_dimensions['K'].width=25
            ws.column_dimensions['L'].width=22
            ws.column_dimensions['M'].width=50

            titulo = NamedStyle(name="titulo")
            titulo.font=Font(size=12,bold=True)
            titulo.fill=PatternFill(fill_type='solid',start_color='00CCCCFF')
            titulo.alignment=Alignment(horizontal='center',mergeCell=True)
            titulo.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )

            celdas = NamedStyle(name="celdas")
            celdas.font=Font(size=12)
            
            celdas.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            celdas_beca = NamedStyle(name="celdas_beca")
            celdas_beca.font=Font(size=12,color='00CCCC')
            
            celdas_beca.alignment=Alignment(horizontal='general',mergeCell=True)
            celdas_beca.border = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                 diagonal=Side(border_style='thin',
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style='thin',
                              color='FF000000'),
                 vertical=Side(border_style='thin',
                               color='FF000000'),
                 horizontal=Side(border_style='thin',
                                color='FF000000')
                )
            #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
            ws['A1'] = 'Usuarios que se el asignaron Becas en el Congresos :'
            ws['A1'].font = Font(size=12,bold=True)
            ws['A1'].alignment = Alignment(mergeCell='center',horizontal='center') 
            
            ws['A2'] ='" %s "'%(query[0]['congreso__titulo']) 
            ws['A2'].font = Font(size=12,bold=True)
            ws['A2'].alignment = Alignment(mergeCell='center',horizontal='center') 
           
            #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
            ws.merge_cells('A1:F1')
            ws.merge_cells('A2:F2')
            #Creamos los encabezados desde la celda B3 hasta la E3
            ws['A3'].style =titulo
            ws['B3'].style =titulo 
            ws['C3'].style =titulo 
            ws['D3'].style =titulo                   
            ws['E3'].style =titulo 
            ws['F3'].style =titulo 
            ws['G3'].style =titulo 
            ws['H3'].style =titulo
            ws['I3'].style =titulo 
            ws['J3'].style =titulo  
            ws['K3'].style =titulo 
            ws['L3'].style =titulo
            ws['M3'].style =titulo


            ws['A3'] = 'No.'
            ws['B3'] = 'Nombre'
            ws['C3'] = 'Email'
            ws['D3'] = 'Dirección'
            ws['E3'] = 'Teléfono'
            ws['F3'] = 'Género'
            ws['G3'] = 'Categoría'
            ws['H3'] = 'Especialidad'
            ws['I3'] = 'Cédula Profecional' 
            ws['J3'] = 'Fecha de Nacimiento' 
            ws['K3'] = 'Categoría de Pago'  
            ws['L3'] = 'Cantidad Comprados' 
            ws['M3'] = 'Lugar de Trabajo'            
            cont=4
            
            #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
            for quer in query:
                
                ws.cell(row=cont,column=1).style=celdas
                ws.cell(row=cont,column=1).value = cont-3
                ws.cell(row=cont,column=2).style=celdas
                ws.cell(row=cont,column=2).value ='%s %s'%(quer['user__usuario__first_name'],quer['user__usuario__last_name']) 
                ws.cell(row=cont,column=3).style=celdas
                ws.cell(row=cont,column=3).value = quer['user__usuario__email']
                ws.cell(row=cont,column=4).style=celdas
                ws.cell(row=cont,column=4).value = quer['user__ubicacion__direccion']
                ws.cell(row=cont,column=5).style=celdas
                ws.cell(row=cont,column=5).value = quer['user__num_telefono']
                ws.cell(row=cont,column=6).style=celdas
                ws.cell(row=cont,column=6).value = quer['user__genero__denominacion']
                ws.cell(row=cont,column=7).style=celdas
                ws.cell(row=cont,column=7).value = quer['user__categoria__nombre']
                ws.cell(row=cont,column=8).style=celdas
                ws.cell(row=cont,column=8).value = quer['user__especialidad__nombre']
                ws.cell(row=cont,column=9).style=celdas
                ws.cell(row=cont,column=9).value = quer['user__cel_profecional']
                ws.cell(row=cont,column=10).style=celdas
                ws.cell(row=cont,column=10).value = quer['user__fecha_nacimiento']
                ws.cell(row=cont,column=11).style=celdas
                ws.cell(row=cont,column=11).value = quer['categoria_pago__nombre']
                ws.cell(row=cont,column=12).style=celdas
                ws.cell(row=cont,column=12).value = quer['cantidad__sum']
                ws.cell(row=cont,column=13).style=celdas
                ws.cell(row=cont,column=13).value = quer['user__puesto']
                cont = cont + 1
                
            
            response = HttpResponse(content_type="application/ms-excel") 
            response["Content-Disposition"] = "attachment; filename=BecasCongreso.xlsx"
            wb.save(response)
            return response
        else:
            congreso=Congreso.objects.get(pk=id_congreso)
            messages.warning(self.request, 'Todavía ningún usuario ha comprado este congreso')
            return HttpResponseRedirect(reverse_lazy('MedCongressAdmin:asig_becas_list')+'?exportar=%s'%(congreso.path))
    def get_queryset(self):
        queryset=RelCongresoUser.objects.filter(is_beca=True)
        return queryset
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs(*args, **kwargs)
        kwargs['user'] =self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(ExportarBecas, self).get_context_data(**kwargs)
        if self.request.user.is_staff :
           
            congresos= Congreso.objects.all()
        else:
            congresos=[]
            org_congresos=Organizador.objects.filter(user=self.request.user.perfilusuario)
            for org_congreso in org_congresos:
                congresos.append(org_congreso.congreso)

        if self.request.GET.get('exportar'):
            congreso_evn=[]
           
            activo=False
            for congreso in congresos:
                if congreso.path == self.request.GET.get('exportar'):
                    activo=True
                else:
                    activo=False
               
                congreso_evn.append({ 'id':congreso.pk,
                                    'titulo':congreso.titulo,
                                        'activo':activo,

                })
            context['exportar']= congreso_evn
        
        context['congresos']= congresos
            
        return context  
 
